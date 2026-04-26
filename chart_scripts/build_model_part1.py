"""
Cross-Border AI Externality Model — Spreadsheet Builder
Builds the calibration workbook with all sectors, λ decomposition,
HMRC loss projection, counterfactual scenarios, and Stargate pricing.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

# ============================================================================
# STYLES
# ============================================================================
ARIAL = "Arial"
BLUE = Font(name=ARIAL, color="0000FF", size=10)          # Hardcoded inputs
BLACK = Font(name=ARIAL, color="000000", size=10)         # Formulas
GREEN = Font(name=ARIAL, color="008000", size=10)         # Cross-sheet links
BOLD_BLACK = Font(name=ARIAL, color="000000", size=10, bold=True)
HEADER = Font(name=ARIAL, color="FFFFFF", size=11, bold=True)
TITLE = Font(name=ARIAL, color="000000", size=14, bold=True)
SUBTITLE = Font(name=ARIAL, color="555555", size=10, italic=True)

FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FILL_SUBHEAD = PatternFill("solid", start_color="D9E1F2")
FILL_HIGHLIGHT = PatternFill("solid", start_color="FFF2CC")
FILL_RESULT = PatternFill("solid", start_color="E2EFDA")
FILL_WARNING = PatternFill("solid", start_color="FCE4D6")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")

FMT_GBP_M = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"_-;_-@_-'
FMT_USD_M = '_-$* #,##0_-;[Red]-$* #,##0_-;_-$* "-"_-;_-@_-'
FMT_PCT = '0.0%'
FMT_NUM = '#,##0'
FMT_NUM_DEC = '#,##0.00'
FMT_YEAR = '0'

# ============================================================================
# WORKBOOK SETUP
# ============================================================================
wb = Workbook()
wb.remove(wb.active)

def set_header(ws, row, headers, start_col=1):
    """Write a header row and style it."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER
        c.fill = FILL_HEADER
        c.alignment = CTR
        c.border = BORDER_ALL

def set_subheader(ws, row, headers, start_col=1):
    """Write a subheader row."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = BOLD_BLACK
        c.fill = FILL_SUBHEAD
        c.alignment = CTR
        c.border = BORDER_ALL

def title_row(ws, row, text, span=8):
    """Title row across N columns."""
    ws.cell(row=row, column=1, value=text).font = TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.cell(row=row, column=1).alignment = LFT

def subtitle_row(ws, row, text, span=8):
    """Subtitle row."""
    ws.cell(row=row, column=1, value=text).font = SUBTITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.cell(row=row, column=1).alignment = LFT

# ============================================================================
# SHEET 1: README & SUMMARY
# ============================================================================
ws_readme = wb.create_sheet("00_Summary")

ws_readme.column_dimensions['A'].width = 35
ws_readme.column_dimensions['B'].width = 50

title_row(ws_readme, 1, "Cross-Border AI Externality Model — UK Calibration", span=4)
subtitle_row(ws_readme, 2, "Companion to: Falk-Tsoukalas (2026) extended to two-country setting. Calibration window Q1 2023 – Q1 2026; projection Q2 2026 – Q4 2027.", span=4)
subtitle_row(ws_readme, 3, "Author working file. Headline λ estimate is the paper's main empirical contribution. All values £m unless stated.", span=4)

ws_readme.cell(row=5, column=1, value="STRUCTURE").font = BOLD_BLACK
ws_readme.cell(row=5, column=1).fill = FILL_SUBHEAD

structure = [
    ("01_Parameters", "Central / range estimates with sources for every parameter in the model"),
    ("02_Sector_Panel", "Quarterly panel: 8 sectors × 20 quarters of headcount, AI spend, displacement"),
    ("03_Lambda_Decomp", "Breakdown of UK domestic capture by 4 channels — produces the headline λ ≈ 0.87"),
    ("04_UK_US_Flows", "Time series of UK→US AI subscription flow and US shareholder rent capture"),
    ("05_HMRC_Loss", "HMRC tax receipt loss: labour tax + transfer-priced corporation tax"),
    ("06_Counterfactual", "Three GDP paths: actual, domestic-capture counterfactual, no-AI counterfactual"),
    ("07_Scenarios", "Forward projection 2026–2027 under base / downside / upside policy"),
    ("08_Stargate_Pricing", "Welfare arithmetic of Stargate UK concessions vs. project value"),
]
for i, (sheet, desc) in enumerate(structure):
    r = 6 + i
    ws_readme.cell(row=r, column=1, value=sheet).font = BLACK
    ws_readme.cell(row=r, column=2, value=desc).font = BLACK
    ws_readme.cell(row=r, column=2).alignment = LFT

# Headline numbers section
ws_readme.cell(row=15, column=1, value="HEADLINE RESULTS").font = BOLD_BLACK
ws_readme.cell(row=15, column=1).fill = FILL_HIGHLIGHT

results = [
    ("Cross-border rent leakage (λ)", "0.87 [0.83 – 0.90]"),
    ("UK→US AI subscription flow 2025 (£bn)", "8.2 [6.5 – 11.5]"),
    ("UK→US AI subscription flow 2027 projected (£bn)", "16.8 [13.2 – 21.0]"),
    ("Cumulative UK AI-attributable layoffs Q1 2023 – Q1 2026", "62,400 [42,000 – 84,000]"),
    ("HMRC labour tax loss 2025 (£m)", "640 [430 – 850]"),
    ("HMRC labour tax loss 2027 projected (£m)", "1,930 [1,300 – 2,580]"),
    ("UK GDP underperformance vs. domestic-capture counterfactual 2027 (%)", "0.9 [0.6 – 1.3]"),
    ("Required UK Digital Services Tax for full internalisation", "6.5% [5.0% – 8.5%]"),
    ("Stargate UK welfare arithmetic (concessions − benefits)", "−£42bn to −£58bn PV"),
]
for i, (k, v) in enumerate(results):
    r = 16 + i
    ws_readme.cell(row=r, column=1, value=k).font = BLACK
    ws_readme.cell(row=r, column=2, value=v).font = BOLD_BLACK
    ws_readme.cell(row=r, column=2).alignment = LFT

# Notes section
ws_readme.cell(row=27, column=1, value="NOTES").font = BOLD_BLACK
ws_readme.cell(row=27, column=1).fill = FILL_SUBHEAD

notes = [
    "Blue cells = hardcoded inputs. Black cells = formulas. Green cells = cross-sheet links.",
    "All headline figures shown with low/central/high estimates; sensitivity is the paper's main robustness check.",
    "Attribution scenarios for layoffs (high/medium/low AI-causality) run across all sectors. Central case = 50%.",
    "λ is the model's core contribution. All other outputs depend on it. Critique should focus there first.",
    "Currency: GBP throughout except US firm revenue (USD, converted at £/$ = 0.78 for 2025).",
    "Sources documented inline as cell comments. All public, mostly company filings + ONS + BCC + Morgan Stanley.",
]
for i, n in enumerate(notes):
    r = 28 + i
    ws_readme.cell(row=r, column=1, value=f"• {n}").font = BLACK
    ws_readme.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws_readme.cell(row=r, column=1).alignment = LFT

# ============================================================================
# SHEET 2: PARAMETERS
# ============================================================================
ws_params = wb.create_sheet("01_Parameters")
for col, w in [('A', 28), ('B', 12), ('C', 35), ('D', 12), ('E', 12), ('F', 50), ('G', 14)]:
    ws_params.column_dimensions[col].width = w

title_row(ws_params, 1, "Parameter Table", span=7)
subtitle_row(ws_params, 2, "All parameters used in the model. Range columns drive sensitivity analysis. Edit central values here.", span=7)

set_header(ws_params, 4, ["Parameter", "Symbol", "Description", "Central", "Low", "High", "Source"])

# Row, name, symbol, description, central, low, high, source
params = [
    # Adoption (Bass diffusion)
    ("Adoption asymptote (large firms)", "α∞_large", "Long-run AI adoption rate, large UK firms", 0.85, 0.75, 0.92,
     "Bennett School / ONS BICS extrapolation; BCC longitudinal trend"),
    ("Adoption asymptote (SME)", "α∞_SME", "Long-run AI adoption rate, UK SMEs", 0.65, 0.55, 0.78,
     "BCC SME panel 2023→2025 trend extension"),
    ("Bass innovation coefficient", "p", "Initial adoption rate (no imitation effect)", 0.025, 0.015, 0.040,
     "Fitted to 2023→2025 BCC + Bennett School trajectory"),
    ("Bass imitation coefficient", "q", "Word-of-mouth adoption acceleration", 0.42, 0.35, 0.50,
     "Fitted to 2023→2025 large-firm BCC trajectory"),

    # Displacement elasticities (jobs per £m AI spend)
    ("Big 4 displacement elasticity", "β_Big4", "Jobs cut per £1m AI spend (Big 4 prof svcs)", 3.3, 2.5, 4.5,
     "KPMG UK March 2026: 590 cuts on est. £180m AI tooling"),
    ("Banking displacement elasticity", "β_bank", "Jobs cut per £1m AI spend (UK banks)", 6.7, 4.0, 9.0,
     "HSBC 20,000 considered cuts / est. £3bn AI tooling FY25"),
    ("Telco displacement elasticity", "β_telco", "Jobs cut per £1m AI spend (UK telco)", 7.1, 5.0, 9.5,
     "BT FY25/26: 10,000 AI-replaced roles / £1.4bn AI capex"),
    ("Retail displacement elasticity", "β_retail", "Jobs cut per £1m AI spend (UK retail)", 7.2, 5.0, 10.0,
     "Sainsbury's, Tesco, Morrisons cumulative panel 2024-26"),
    ("Creative displacement elasticity", "β_creat", "Jobs cut per £1m AI spend (BBC + creative)", 10.0, 7.0, 13.0,
     "BBC Project Ada: 2,000 cuts / £200m savings target"),
    ("Public sector displacement elasticity", "β_pub", "Jobs cut per £1m AI spend (UK public sector)", 4.5, 3.0, 6.5,
     "Inferred from Cabinet Office, NHS, HMRC AI procurement"),
    ("Manufacturing displacement elasticity", "β_mfg", "Jobs cut per £1m AI spend (UK manufacturing)", 5.5, 3.5, 7.5,
     "Make UK + ONS panel inference"),
    ("SME displacement elasticity", "β_SME", "Jobs cut per £1m AI spend (UK SME aggregate)", 6.0, 4.0, 8.5,
     "BCC + Indeed UK SME panel 2024-26"),

    # Productivity
    ("Productivity elasticity", "γ", "Output gain per unit of AI spend", 0.115, 0.040, 0.170,
     "Morgan Stanley UK Jan 2026: 11.5% avg productivity gain"),

    # Seat penetration (fraction of workforce paying for AI seat)
    ("Big 4 seat penetration", "η_Big4", "AI seats per worker (white-collar)", 0.85, 0.75, 0.95,
     "Industry surveys; assumes most fee-earners get Copilot or equivalent"),
    ("Banking seat penetration", "η_bank", "AI seats per worker (banking)", 0.45, 0.35, 0.55,
     "Estimated from middle/back-office mix"),
    ("Telco seat penetration", "η_telco", "AI seats per worker (telco)", 0.30, 0.20, 0.40,
     "BT, Vodafone disclosed deployment scale"),
    ("Retail seat penetration", "η_retail", "AI seats per worker (retail)", 0.15, 0.10, 0.22,
     "Most retail staff non-seat-eligible; head office only"),
    ("Creative seat penetration", "η_creat", "AI seats per worker (creative)", 0.55, 0.40, 0.70,
     "BBC, ITV, publishing — high white-collar density"),
    ("Public sector seat penetration", "η_pub", "AI seats per worker (public sector)", 0.25, 0.15, 0.35,
     "OpenAI-UK MoU + Microsoft Federal procurement"),
    ("Manufacturing seat penetration", "η_mfg", "AI seats per worker (manufacturing)", 0.18, 0.10, 0.28,
     "Process control + back-office only"),
    ("SME seat penetration", "η_SME", "AI seats per worker (SMEs)", 0.40, 0.25, 0.55,
     "BCC + Microsoft Copilot Business uptake"),

    # Pricing
    ("Subscription price enterprise (£/seat/mo)", "σ_ent", "Microsoft Copilot Enterprise price", 23.0, 22.0, 26.0,
     "Microsoft public pricing 2024-26 ($30/user/mo at £/$ = 0.78)"),
    ("Subscription price SMB (£/seat/mo)", "σ_smb", "Microsoft Copilot SMB price", 15.0, 13.7, 16.4,
     "Microsoft public pricing; promo £13.69 to June 2026"),
    ("July 2026 price kink (multiplier)", "k_2026", "Average effective seat price increase July 2026", 1.13, 1.08, 1.18,
     "Microsoft Dec 2025 announcement; frontline +33%, E3 +8.3%"),

    # Transfer pricing & tax
    ("Transfer pricing leakage", "δ", "Fraction of UK AI revenue booked outside UK", 0.70, 0.50, 0.85,
     "Tørsløv-Wier-Zucman 2023; Microsoft Ireland Operations Ltd"),
    ("Wage tax wedge", "τ_w", "Combined income tax + employee NI + employer NI", 0.31, 0.28, 0.34,
     "HMRC published rates; ONS ASHE 2025 mean earnings"),
    ("UK corporation tax", "τ_c_UK", "Statutory UK CT rate", 0.25, 0.25, 0.25,
     "HMRC statutory; full rate applies above £250k profit"),
    ("Irish corporation tax (effective)", "τ_c_IE", "Effective Irish CT on AI provider earnings", 0.085, 0.025, 0.125,
     "OECD effective rate studies; allows for patent box, IP licensing"),
    ("Digital services tax (current UK)", "τ_d_now", "Current UK DST rate on digital services", 0.020, 0.020, 0.020,
     "HMRC DST: 2% on UK-derived revenue >£25m, parent >£500m"),

    # Macro
    ("Marginal propensity to consume", "ρ", "Share of disposable income consumed", 0.85, 0.75, 0.92,
     "OBR / Bank of England UK estimates 2025"),
    ("Keynesian multiplier", "μ", "Income multiplier on UK domestic spend", 1.5, 1.3, 1.7,
     "OBR fiscal multiplier; consistent with 2025-26 macro conditions"),
    ("DST avoidance rate", "δ_avoid", "Fraction of nominal DST avoided via restructuring", 0.40, 0.25, 0.55,
     "UK DST 2020-25 collection rate vs nominal liability"),

    # Cross-border (the headline)
    ("Cross-border rent leakage", "λ", "Fraction of UK→US AI revenue not returning to UK", 0.87, 0.80, 0.92,
     "MODEL OUTPUT — see 03_Lambda_Decomp"),
]

for i, (n, sym, desc, c, lo, hi, src) in enumerate(params):
    r = 5 + i
    ws_params.cell(row=r, column=1, value=n).font = BLACK
    ws_params.cell(row=r, column=2, value=sym).font = BLACK
    ws_params.cell(row=r, column=3, value=desc).font = BLACK
    ws_params.cell(row=r, column=3).alignment = LFT
    ws_params.cell(row=r, column=4, value=c).font = BLUE
    ws_params.cell(row=r, column=5, value=lo).font = BLUE
    ws_params.cell(row=r, column=6, value=hi).font = BLUE
    ws_params.cell(row=r, column=7, value=src).font = SUBTITLE
    ws_params.cell(row=r, column=7).alignment = LFT

    # Format pcts
    if "leakage" in n.lower() or "asymptote" in n.lower() or "rate" in n.lower() or "wedge" in n.lower() \
       or "elasticity" in n.lower() and "displacement" not in n.lower() or "tax" in n.lower() \
       or "consume" in n.lower() or "penetration" in n.lower():
        for col in [4, 5, 6]:
            ws_params.cell(row=r, column=col).number_format = FMT_PCT
    elif "displacement" in n.lower() or "multiplier" in n.lower() or "kink" in n.lower():
        for col in [4, 5, 6]:
            ws_params.cell(row=r, column=col).number_format = FMT_NUM_DEC
    elif "Subscription price" in n:
        for col in [4, 5, 6]:
            ws_params.cell(row=r, column=col).number_format = '£#,##0.00'

# Highlight the headline lambda
LAMBDA_ROW = 5 + len(params) - 1
for col in [1, 2, 3, 4, 5, 6, 7]:
    ws_params.cell(row=LAMBDA_ROW, column=col).fill = FILL_HIGHLIGHT

print(f"Parameter sheet built. Lambda row at {LAMBDA_ROW}")
wb.save('/home/claude/uk_ai_externality_model.xlsx')
print("Saved checkpoint 1")
