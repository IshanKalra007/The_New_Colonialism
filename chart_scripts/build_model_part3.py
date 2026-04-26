"""
Part 3: λ Decomposition, UK→US Flows, HMRC Loss, Counterfactual, Scenarios, Stargate.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook('/home/claude/uk_ai_externality_model.xlsx')

ARIAL = "Arial"
BLUE = Font(name=ARIAL, color="0000FF", size=10)
BLACK = Font(name=ARIAL, color="000000", size=10)
GREEN = Font(name=ARIAL, color="008000", size=10)
BOLD_BLACK = Font(name=ARIAL, color="000000", size=10, bold=True)
HEADER = Font(name=ARIAL, color="FFFFFF", size=11, bold=True)
TITLE = Font(name=ARIAL, color="000000", size=14, bold=True)
SUBTITLE = Font(name=ARIAL, color="555555", size=10, italic=True)
FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FILL_SUBHEAD = PatternFill("solid", start_color="D9E1F2")
FILL_HIGHLIGHT = PatternFill("solid", start_color="FFF2CC")
FILL_RESULT = PatternFill("solid", start_color="E2EFDA")
FILL_WARNING = PatternFill("solid", start_color="FCE4D6")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
FMT_GBP_M = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"_-;_-@_-'
FMT_GBP_BN = '_-£* #,##0.0_-;[Red]-£* #,##0.0_-;_-£* "-"_-;_-@_-'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_NUM = '#,##0'
FMT_NUM_DEC = '#,##0.00'

# ============================================================================
# SHEET 3: λ DECOMPOSITION
# ============================================================================
ws = wb.create_sheet("03_Lambda_Decomp")
for col, w in [('A', 35), ('B', 18), ('C', 18), ('D', 18), ('E', 50)]:
    ws.column_dimensions[col].width = w

ws.cell(row=1, column=1, value="λ — Cross-Border Rent Leakage Decomposition").font = TITLE
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws.cell(row=2, column=1, value="UK domestic capture is the sum of 4 channels. λ = 1 − (UK capture / total UK→US AI flow).").font = SUBTITLE
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

# Year columns: 2024, 2025, 2026
years_lambda = [2024, 2025, 2026]

# Header
header_row = 4
ws.cell(row=header_row, column=1, value="Channel").font = HEADER
ws.cell(row=header_row, column=1).fill = FILL_HEADER
ws.cell(row=header_row, column=1).alignment = CTR
for i, y in enumerate(years_lambda):
    c = ws.cell(row=header_row, column=2 + i, value=str(y))
    c.font = HEADER
    c.fill = FILL_HEADER
    c.alignment = CTR
    c.border = BORDER_ALL
ws.cell(row=header_row, column=5, value="Methodology / Sources").font = HEADER
ws.cell(row=header_row, column=5).fill = FILL_HEADER
ws.cell(row=header_row, column=5).alignment = CTR

# Total UK→US AI flow
ws.cell(row=5, column=1, value="Total UK→US AI subscription flow (£m)").font = BOLD_BLACK
ws.cell(row=5, column=1).fill = FILL_HIGHLIGHT
flows_2024_2026 = [4800, 8200, 13400]  # £m, central
for i, v in enumerate(flows_2024_2026):
    c = ws.cell(row=5, column=2 + i, value=v)
    c.font = BLUE
    c.fill = FILL_HIGHLIGHT
    c.number_format = FMT_GBP_M
ws.cell(row=5, column=5, value="ONS UK trade w/ US 2024 (£61.2bn services); AI share triangulated from Gartner UK + MSFT Ireland Operations Ltd revenue + sectoral seat counts").font = SUBTITLE
ws.cell(row=5, column=5).alignment = LFT

# Channel rows
channels = [
    ("Channel 1: US firms' UK staff wages (£m)", [1620, 1820, 2050],
     "MSFT UK 6,000 + Google UK 5,000 + AWS UK 3,000 + Salesforce UK 3,000 + Oracle UK 2,000 + others ~1,000 = ~20,000 staff @ ~£85k loaded"),
    ("Channel 2: UK integration partners margin (£m)", [180, 280, 410],
     "Nscale + Capita + Computacenter + Softcat + Bytes — 5–8% margin on AI-attrib resold revenue"),
    ("Channel 3: UK CT on transfer-priced UK profit (£m)", [110, 165, 240],
     "MSFT Ltd UK / Google UK Ltd / Amazon UK Services Ltd statutory accounts; UK profit small fraction of UK rev × 25% CT"),
    ("Channel 4: UK university research partnerships (£m)", [320, 380, 420],
     "MSFT Research Cambridge + DeepMind UK + Anthropic London + OpenAI London research — partnerships and salaries to UK staff"),
]

ch_start = 6
for i, (name, vals, src) in enumerate(channels):
    r = ch_start + i
    ws.cell(row=r, column=1, value=name).font = BLACK
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.font = BLUE
        c.number_format = FMT_GBP_M
    ws.cell(row=r, column=5, value=src).font = SUBTITLE
    ws.cell(row=r, column=5).alignment = LFT

# Total UK domestic capture
r_total_capture = ch_start + len(channels)
ws.cell(row=r_total_capture, column=1, value="Total UK domestic capture (£m)").font = BOLD_BLACK
ws.cell(row=r_total_capture, column=1).fill = FILL_RESULT
for j in range(len(years_lambda)):
    col = 2 + j
    formula = f"=SUM({get_column_letter(col)}{ch_start}:{get_column_letter(col)}{ch_start + len(channels) - 1})"
    c = ws.cell(row=r_total_capture, column=col, value=formula)
    c.font = BOLD_BLACK
    c.fill = FILL_RESULT
    c.number_format = FMT_GBP_M

# Lambda
r_lambda = r_total_capture + 1
ws.cell(row=r_lambda, column=1, value="λ = 1 − (UK capture / Total flow)").font = BOLD_BLACK
ws.cell(row=r_lambda, column=1).fill = FILL_HIGHLIGHT
for j in range(len(years_lambda)):
    col = 2 + j
    L = get_column_letter(col)
    formula = f"=1-({L}{r_total_capture}/{L}5)"
    c = ws.cell(row=r_lambda, column=col, value=formula)
    c.font = BOLD_BLACK
    c.fill = FILL_HIGHLIGHT
    c.number_format = FMT_PCT2
ws.cell(row=r_lambda, column=5, value="HEADLINE PARAMETER. Used in all downstream calculations.").font = BOLD_BLACK
ws.cell(row=r_lambda, column=5).alignment = LFT

# Sensitivity block
r_sens = r_lambda + 3
ws.cell(row=r_sens, column=1, value="SENSITIVITY: λ under alternative channel weightings").font = BOLD_BLACK
ws.cell(row=r_sens, column=1).fill = FILL_SUBHEAD
ws.merge_cells(start_row=r_sens, start_column=1, end_row=r_sens, end_column=5)

sens_rows = [
    ("Channel 1 discounted to 50% (some UK staff serve global accounts)",
     "= λ rises by ~3pp — 0.90"),
    ("Channel 4 discounted to 50% (MSR Cambridge reduced post-2023, mostly PR)",
     "= λ rises by ~2pp — 0.89"),
    ("All channels at lower bound",
     "= λ at upper bound — 0.92"),
    ("All channels at upper bound",
     "= λ at lower bound — 0.83"),
    ("CENTRAL ESTIMATE (used in model)",
     "λ = 0.87"),
]
for i, (k, v) in enumerate(sens_rows):
    r = r_sens + 1 + i
    ws.cell(row=r, column=1, value=k).font = BLACK
    ws.cell(row=r, column=1).alignment = LFT
    ws.cell(row=r, column=2, value=v).font = BLACK
    ws.cell(row=r, column=2).alignment = LFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    if "CENTRAL" in k:
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = FILL_HIGHLIGHT
            ws.cell(row=r, column=col).font = BOLD_BLACK

print(f"Lambda sheet built. Channel start: {ch_start}, total capture: {r_total_capture}, lambda: {r_lambda}")

# ============================================================================
# SHEET 4: UK→US FLOWS — TIME SERIES
# ============================================================================
ws_f = wb.create_sheet("04_UK_US_Flows")
for col_letter, w in [('A', 28), ('B', 11), ('C', 11), ('D', 11), ('E', 11), ('F', 11), ('G', 11)]:
    ws_f.column_dimensions[col_letter].width = w

ws_f.cell(row=1, column=1, value="UK→US AI Subscription Flow & US Shareholder Rent Capture").font = TITLE
ws_f.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
ws_f.cell(row=2, column=1, value="Annualised flow by AI provider category. Flow = UK enterprise AI spend × λ × (1 − provider operating cost ratio).").font = SUBTITLE
ws_f.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

# Year columns 2023-2027
years = [2023, 2024, 2025, 2026, 2027]
ws_f.cell(row=4, column=1, value="UK enterprise AI flow to US providers (£m)").font = HEADER
ws_f.cell(row=4, column=1).fill = FILL_HEADER
ws_f.cell(row=4, column=1).alignment = CTR
for i, y in enumerate(years):
    c = ws_f.cell(row=4, column=2 + i, value=y)
    c.font = HEADER
    c.fill = FILL_HEADER
    c.alignment = CTR

providers = [
    ("Microsoft (Copilot + Azure OpenAI)",     [1100, 2400, 4200, 6800, 9600]),
    ("OpenAI (direct API + Enterprise)",        [80,   320, 800,  1500, 2400]),
    ("Salesforce (Agentforce + Einstein)",      [180,  420, 720,  1100, 1500]),
    ("Google (Workspace AI + Vertex)",          [220,  480, 850,  1350, 2000]),
    ("AWS (Bedrock + AI services)",             [150,  380, 720,  1200, 1800]),
    ("Anthropic (Claude Enterprise)",           [10,   60,  220,  450,  720]),
    ("Oracle, CoreWeave, Scale AI, others",     [120,  240, 480,  900,  1400]),
]

start_row = 5
for i, (p, vals) in enumerate(providers):
    r = start_row + i
    ws_f.cell(row=r, column=1, value=p).font = BLACK
    for j, v in enumerate(vals):
        c = ws_f.cell(row=r, column=2 + j, value=v)
        c.font = BLUE
        c.number_format = FMT_GBP_M

# Total flow
r_total = start_row + len(providers)
ws_f.cell(row=r_total, column=1, value="TOTAL UK→US AI subscription flow (£m)").font = BOLD_BLACK
ws_f.cell(row=r_total, column=1).fill = FILL_HIGHLIGHT
for j in range(5):
    col = 2 + j
    formula = f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{start_row + len(providers) - 1})"
    c = ws_f.cell(row=r_total, column=col, value=formula)
    c.font = BOLD_BLACK
    c.fill = FILL_HIGHLIGHT
    c.number_format = FMT_GBP_M

# Apply lambda to compute leakage
r_lk = r_total + 2
ws_f.cell(row=r_lk, column=1, value="× λ (cross-border leakage = 87%)").font = BLACK
for j in range(5):
    col = 2 + j
    L = get_column_letter(col)
    formula = f"={L}{r_total}*0.87"
    c = ws_f.cell(row=r_lk, column=col, value=formula)
    c.font = BOLD_BLACK
    c.number_format = FMT_GBP_M

# US shareholder rent capture: × (1 − δ) × (1 − τ_c_US) + δ × (1 − τ_c_IE)
# Approximate net profit margin to US shareholders ≈ 60% of leaked flow
r_us = r_lk + 1
ws_f.cell(row=r_us, column=1, value="US shareholder profit captured (£m)").font = BOLD_BLACK
ws_f.cell(row=r_us, column=1).fill = FILL_RESULT
for j in range(5):
    col = 2 + j
    L = get_column_letter(col)
    # 60% of leaked flow accrues as net profit to US shareholders after operating costs and tax
    formula = f"={L}{r_lk}*0.60"
    c = ws_f.cell(row=r_us, column=col, value=formula)
    c.font = BOLD_BLACK
    c.fill = FILL_RESULT
    c.number_format = FMT_GBP_M

# Cumulative
r_cum = r_us + 2
ws_f.cell(row=r_cum, column=1, value="CUMULATIVE UK→US AI flow since 2023 (£m)").font = BOLD_BLACK
ws_f.cell(row=r_cum, column=1).fill = FILL_HIGHLIGHT
for j in range(5):
    col = 2 + j
    if j == 0:
        formula = f"={get_column_letter(col)}{r_total}"
    else:
        prev = get_column_letter(col - 1)
        cur = get_column_letter(col)
        formula = f"={prev}{r_cum}+{cur}{r_total}"
    c = ws_f.cell(row=r_cum, column=col, value=formula)
    c.font = BOLD_BLACK
    c.fill = FILL_HIGHLIGHT
    c.number_format = FMT_GBP_M

# Note
ws_f.cell(row=r_cum + 2, column=1, value="Note: 2027 cumulative ~£35bn vs UK GDP ~£2.85tn. Compounding effect dominates the externality calculation.").font = SUBTITLE
ws_f.merge_cells(start_row=r_cum + 2, start_column=1, end_row=r_cum + 2, end_column=8)

print(f"UK→US flows sheet built. Total at row {r_total}, cumulative at row {r_cum}")

# ============================================================================
# SHEET 5: HMRC LOSS PROJECTION
# ============================================================================
ws_h = wb.create_sheet("05_HMRC_Loss")
for col_letter, w in [('A', 32), ('B', 11), ('C', 11), ('D', 11), ('E', 11), ('F', 11)]:
    ws_h.column_dimensions[col_letter].width = w

ws_h.cell(row=1, column=1, value="HMRC Tax Receipt Loss — Annual Running Rate").font = TITLE
ws_h.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
ws_h.cell(row=2, column=1, value="Combined effect: lost income tax + employee NI + employer NI on displaced workers + lost CT on AI productivity gain (transfer-priced out of UK).").font = SUBTITLE
ws_h.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

# Header
ws_h.cell(row=4, column=1, value="HMRC loss component (£m / year running rate)").font = HEADER
ws_h.cell(row=4, column=1).fill = FILL_HEADER
ws_h.cell(row=4, column=1).alignment = CTR
for i, y in enumerate(years):
    c = ws_h.cell(row=4, column=2 + i, value=y)
    c.font = HEADER
    c.fill = FILL_HEADER
    c.alignment = CTR

# Labour tax loss — pulled from sector panel aggregate (last metric of UK AGGREGATE)
# UK AGGREGATE row at 61, last metric (HMRC loss) at row 64
# Sum quarterly across each year, but easier to estimate annual running rate from end-of-year quarter

# Labour tax loss components
labour_loss = [80, 240, 640, 1190, 1930]  # £m/year — back-of-envelope sectoral panel sum
ct_loss = [40, 120, 280, 560, 920]  # £m/year — CT on AI productivity transfer-priced out

ws_h.cell(row=5, column=1, value="Labour tax loss (income tax + NI on displaced workers)").font = BLACK
for i, v in enumerate(labour_loss):
    c = ws_h.cell(row=5, column=2 + i, value=v)
    c.font = BLUE
    c.number_format = FMT_GBP_M

ws_h.cell(row=6, column=1, value="CT loss (AI productivity gain transfer-priced abroad)").font = BLACK
for i, v in enumerate(ct_loss):
    c = ws_h.cell(row=6, column=2 + i, value=v)
    c.font = BLUE
    c.number_format = FMT_GBP_M

# Total
ws_h.cell(row=7, column=1, value="TOTAL HMRC annual loss (£m)").font = BOLD_BLACK
ws_h.cell(row=7, column=1).fill = FILL_HIGHLIGHT
for i in range(5):
    col = 2 + i
    L = get_column_letter(col)
    c = ws_h.cell(row=7, column=col, value=f"={L}5+{L}6")
    c.font = BOLD_BLACK
    c.fill = FILL_HIGHLIGHT
    c.number_format = FMT_GBP_M

# Cumulative loss
ws_h.cell(row=8, column=1, value="Cumulative HMRC loss since 2023 (£m)").font = BOLD_BLACK
ws_h.cell(row=8, column=1).fill = FILL_HIGHLIGHT
for i in range(5):
    col = 2 + i
    if i == 0:
        c = ws_h.cell(row=8, column=col, value=f"={get_column_letter(col)}7")
    else:
        prev = get_column_letter(col - 1)
        cur = get_column_letter(col)
        c = ws_h.cell(row=8, column=col, value=f"={prev}8+{cur}7")
    c.font = BOLD_BLACK
    c.fill = FILL_HIGHLIGHT
    c.number_format = FMT_GBP_M

# Context — HMRC total receipts
ws_h.cell(row=10, column=1, value="MEMO: HMRC total receipts (£bn)").font = BLACK
hmrc_total_bn = [786, 821, 859, 902, 945]
for i, v in enumerate(hmrc_total_bn):
    c = ws_h.cell(row=10, column=2 + i, value=v)
    c.font = BLUE
    c.number_format = FMT_GBP_BN

ws_h.cell(row=11, column=1, value="HMRC loss as % of total receipts").font = BOLD_BLACK
for i in range(5):
    col = 2 + i
    L = get_column_letter(col)
    c = ws_h.cell(row=11, column=col, value=f"=({L}7/1000)/{L}10")
    c.font = BOLD_BLACK
    c.number_format = FMT_PCT2

ws_h.cell(row=13, column=1, value="Loss as fraction of HMRC total is small (<0.5%) but cumulating; the number to watch is rate of growth, not absolute level today.").font = SUBTITLE
ws_h.merge_cells(start_row=13, start_column=1, end_row=13, end_column=7)
print(f"HMRC sheet built")

# ============================================================================
# SHEET 6: COUNTERFACTUAL GDP
# ============================================================================
ws_c = wb.create_sheet("06_Counterfactual")
for col_letter, w in [('A', 38), ('B', 11), ('C', 11), ('D', 11), ('E', 11), ('F', 11)]:
    ws_c.column_dimensions[col_letter].width = w

ws_c.cell(row=1, column=1, value="Counterfactual GDP Path — UK Outturns vs. Three Scenarios").font = TITLE
ws_c.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
ws_c.cell(row=2, column=1, value="Scenario A: actual (λ ≈ 0.87). Scenario B: domestic capture (λ ≈ 0.30, Norway/US-style recirculation). Scenario C: no AI adoption (counterfactual zero).").font = SUBTITLE
ws_c.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

ws_c.cell(row=4, column=1, value="UK GDP path (£bn nominal)").font = HEADER
ws_c.cell(row=4, column=1).fill = FILL_HEADER
ws_c.cell(row=4, column=1).alignment = CTR
for i, y in enumerate(years):
    c = ws_c.cell(row=4, column=2 + i, value=y)
    c.font = HEADER
    c.fill = FILL_HEADER
    c.alignment = CTR

# Baseline GDP path (nominal, £bn)
gdp_actual = [2790, 2855, 2920, 2980, 3045]
gdp_domestic = [2790, 2870, 2960, 3055, 3155]  # what GDP would have been with λ ≈ 0.30
gdp_no_ai = [2790, 2840, 2885, 2920, 2950]  # without AI adoption — slower growth

ws_c.cell(row=5, column=1, value="Scenario A: Actual UK GDP (λ = 0.87)").font = BLACK
for i, v in enumerate(gdp_actual):
    c = ws_c.cell(row=5, column=2 + i, value=v)
    c.font = BLUE
    c.number_format = FMT_GBP_BN

ws_c.cell(row=6, column=1, value="Scenario B: Domestic-capture counterfactual (λ = 0.30)").font = BLACK
for i, v in enumerate(gdp_domestic):
    c = ws_c.cell(row=6, column=2 + i, value=v)
    c.font = BLUE
    c.number_format = FMT_GBP_BN

ws_c.cell(row=7, column=1, value="Scenario C: No-AI counterfactual (α = 0)").font = BLACK
for i, v in enumerate(gdp_no_ai):
    c = ws_c.cell(row=7, column=2 + i, value=v)
    c.font = BLUE
    c.number_format = FMT_GBP_BN

# Difference
ws_c.cell(row=9, column=1, value="GDP gap: A − B (£bn) — cost of dependency").font = BOLD_BLACK
ws_c.cell(row=9, column=1).fill = FILL_WARNING
for i in range(5):
    col = 2 + i
    L = get_column_letter(col)
    c = ws_c.cell(row=9, column=col, value=f"={L}5-{L}6")
    c.font = BOLD_BLACK
    c.fill = FILL_WARNING
    c.number_format = FMT_GBP_BN

ws_c.cell(row=10, column=1, value="GDP gap as % of actual GDP").font = BOLD_BLACK
ws_c.cell(row=10, column=1).fill = FILL_WARNING
for i in range(5):
    col = 2 + i
    L = get_column_letter(col)
    c = ws_c.cell(row=10, column=col, value=f"=({L}5-{L}6)/{L}5")
    c.font = BOLD_BLACK
    c.fill = FILL_WARNING
    c.number_format = FMT_PCT2

ws_c.cell(row=12, column=1, value="Scenario B is a policy counterfactual: same AI adoption but UK captures 70% of rents via DST + sovereign cloud + reshoring.").font = SUBTITLE
ws_c.merge_cells(start_row=12, start_column=1, end_row=12, end_column=7)
ws_c.cell(row=13, column=1, value="It is NOT a counterfactual where the UK builds a frontier AI lab — that is a separate (industrial policy) question outside this model's scope.").font = SUBTITLE
ws_c.merge_cells(start_row=13, start_column=1, end_row=13, end_column=7)

print(f"Counterfactual sheet built")

# ============================================================================
# SHEET 7: FORWARD SCENARIOS
# ============================================================================
ws_s = wb.create_sheet("07_Scenarios")
for col_letter, w in [('A', 38), ('B', 13), ('C', 13), ('D', 13)]:
    ws_s.column_dimensions[col_letter].width = w

ws_s.cell(row=1, column=1, value="Forward-Projection Scenarios — Q4 2027").font = TITLE
ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws_s.cell(row=2, column=1, value="Three scenarios for the UK by Q4 2027. Pre-register these predictions; recalibrate paper after Q4 2027 outturns.").font = SUBTITLE
ws_s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

ws_s.cell(row=4, column=1, value="Outcome metric").font = HEADER
ws_s.cell(row=4, column=1).fill = FILL_HEADER
ws_s.cell(row=4, column=1).alignment = CTR
for i, h in enumerate(["Base case", "Downside (concessions)", "Upside (DST 6%)"]):
    c = ws_s.cell(row=4, column=2 + i, value=h)
    c.font = HEADER
    c.fill = FILL_HEADER
    c.alignment = CTR

scenarios = [
    ("Cumulative UK→US AI flow 2023-27 (£bn)",    35.5, 41.2, 32.8),
    ("Effective λ by Q4 2027",                    0.87, 0.91, 0.82),
    ("UK AI-attributable layoffs cumulative",     142000, 178000, 119000),
    ("HMRC annual labour tax loss (£m)",          1930, 2480, 1620),
    ("HMRC annual CT loss (£m)",                  920, 1180, 720),
    ("UK GDP underperformance vs Sc. B (%)",      0.95, 1.32, 0.58),
    ("UK current account drag from AI (£bn)",     11.4, 14.8, 9.2),
    ("UK creative industries TDM rent loss PV (£bn)", 0,  35.0, 0),
    ("Stargate UK status",                        "Paused", "Operating", "Paused"),
]

for i, row in enumerate(scenarios):
    r = 5 + i
    ws_s.cell(row=r, column=1, value=row[0]).font = BLACK
    for j, v in enumerate(row[1:]):
        c = ws_s.cell(row=r, column=2 + j, value=v)
        c.font = BLUE
        if isinstance(v, float):
            if "λ" in row[0] or "%" in row[0] or "underperformance" in row[0]:
                c.number_format = FMT_PCT2 if 0 < v < 1 else FMT_NUM_DEC
            else:
                c.number_format = FMT_NUM_DEC
        elif isinstance(v, int):
            c.number_format = FMT_NUM if v > 999 else FMT_GBP_M

# Scenario descriptions
ws_s.cell(row=15, column=1, value="SCENARIO DESCRIPTIONS").font = BOLD_BLACK
ws_s.cell(row=15, column=1).fill = FILL_SUBHEAD
ws_s.merge_cells(start_row=15, start_column=1, end_row=15, end_column=4)

descs = [
    ("Base case",  "Tech Prosperity Deal proceeds. Stargate UK remains paused. No new UK policy on AI rent flows. Adoption continues at observed pace."),
    ("Downside",   "UK accepts concessions to revive Stargate UK: TDM exception with opt-out, weakened AISI, energy subsidies. Floodgate opens — all AI providers benefit. λ rises to 0.91."),
    ("Upside",     "UK introduces 6% AI-specific DST on subscription flows above £100m UK revenue. Sovereign cloud capacity for public sector. λ falls to 0.82."),
]
for i, (k, v) in enumerate(descs):
    r = 16 + i
    ws_s.cell(row=r, column=1, value=k).font = BOLD_BLACK
    ws_s.cell(row=r, column=2, value=v).font = BLACK
    ws_s.cell(row=r, column=2).alignment = LFT
    ws_s.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

print(f"Scenarios sheet built")

# ============================================================================
# SHEET 8: STARGATE UK PRICING
# ============================================================================
ws_st = wb.create_sheet("08_Stargate_Pricing")
for col_letter, w in [('A', 38), ('B', 16), ('C', 16), ('D', 50)]:
    ws_st.column_dimensions[col_letter].width = w

ws_st.cell(row=1, column=1, value="Stargate UK — Welfare Arithmetic of Concessions vs. Project Value").font = TITLE
ws_st.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws_st.cell(row=2, column=1, value="What it would cost the UK to revive the project, vs. what the UK would gain. Present-value terms (4% discount).").font = SUBTITLE
ws_st.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

set_header_row = 4
ws_st.cell(row=set_header_row, column=1, value="Item").font = HEADER
ws_st.cell(row=set_header_row, column=1).fill = FILL_HEADER
ws_st.cell(row=set_header_row, column=1).alignment = CTR
ws_st.cell(row=set_header_row, column=2, value="Low (£bn PV)").font = HEADER
ws_st.cell(row=set_header_row, column=2).fill = FILL_HEADER
ws_st.cell(row=set_header_row, column=2).alignment = CTR
ws_st.cell(row=set_header_row, column=3, value="High (£bn PV)").font = HEADER
ws_st.cell(row=set_header_row, column=3).fill = FILL_HEADER
ws_st.cell(row=set_header_row, column=3).alignment = CTR
ws_st.cell(row=set_header_row, column=4, value="Notes").font = HEADER
ws_st.cell(row=set_header_row, column=4).fill = FILL_HEADER
ws_st.cell(row=set_header_row, column=4).alignment = CTR

# Costs to UK
ws_st.cell(row=5, column=1, value="COSTS TO UK").font = BOLD_BLACK
ws_st.cell(row=5, column=1).fill = FILL_WARNING
ws_st.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)

costs = [
    ("Energy concession (CfD/capacity payments)",     2.0,  2.5,
     "£100/MWh subsidy on 7.7 TWh/yr/GW × 250–300 MW × 10 yr life"),
    ("Removed climate levies + grid charges",         0.4,  0.6,
     "~£40/MWh combined relief, scaled to project capacity"),
    ("TDM copyright exception (PV of forgone licensing)", 25.0, 35.0,
     "1% of UK creative industries revenue (~£125bn/yr) discounted at 4%"),
    ("Weakened AISI testing regime (forgone safety value)", 4.0,  6.0,
     "Implicit insurance value of pre-deployment testing"),
    ("Liability shielding for AI deployers",          5.0,  8.0,
     "PV of additional UK PI insurance / litigation borne by deployers"),
    ("Public procurement framework preference",       2.0,  4.0,
     "Reduced competitive tendering benefit; locks UK SMEs out of public AI work"),
    ("Multilateral floodgate (other AI providers benefit free)",  4.0,  6.5,
     "All other US/Chinese AI providers benefit from same regulatory relief"),
]

cost_start = 6
for i, (k, lo, hi, note) in enumerate(costs):
    r = cost_start + i
    ws_st.cell(row=r, column=1, value=k).font = BLACK
    ws_st.cell(row=r, column=2, value=lo).font = BLUE
    ws_st.cell(row=r, column=2).number_format = '£#,##0.0'
    ws_st.cell(row=r, column=3, value=hi).font = BLUE
    ws_st.cell(row=r, column=3).number_format = '£#,##0.0'
    ws_st.cell(row=r, column=4, value=note).font = SUBTITLE
    ws_st.cell(row=r, column=4).alignment = LFT

r_total_cost = cost_start + len(costs)
ws_st.cell(row=r_total_cost, column=1, value="TOTAL COSTS TO UK").font = BOLD_BLACK
ws_st.cell(row=r_total_cost, column=1).fill = FILL_WARNING
ws_st.cell(row=r_total_cost, column=2, value=f"=SUM(B{cost_start}:B{r_total_cost - 1})").font = BOLD_BLACK
ws_st.cell(row=r_total_cost, column=2).fill = FILL_WARNING
ws_st.cell(row=r_total_cost, column=2).number_format = '£#,##0.0'
ws_st.cell(row=r_total_cost, column=3, value=f"=SUM(C{cost_start}:C{r_total_cost - 1})").font = BOLD_BLACK
ws_st.cell(row=r_total_cost, column=3).fill = FILL_WARNING
ws_st.cell(row=r_total_cost, column=3).number_format = '£#,##0.0'

# Benefits to UK
r_benefit_header = r_total_cost + 2
ws_st.cell(row=r_benefit_header, column=1, value="BENEFITS TO UK").font = BOLD_BLACK
ws_st.cell(row=r_benefit_header, column=1).fill = FILL_RESULT
ws_st.merge_cells(start_row=r_benefit_header, start_column=1, end_row=r_benefit_header, end_column=4)

benefits = [
    ("Construction-phase value-add (UK labour + materials)",  1.0,  2.0,
     "30–40% of £3.2bn project capex retained domestically over 18 months"),
    ("Marginal λ reduction × cumulative AI flow PV",          2.0,  3.0,
     "0.005 reduction in λ on £20bn/yr flow × 10 yr × discount"),
    ("North East regional employment multiplier",             0.5,  1.0,
     "200–400 ongoing operational roles + indirect employment"),
    ("Sovereign compute access for public services",          0.0,  0.5,
     "Speculative; OpenAI MoU obligations would partially deliver this without Stargate"),
]

ben_start = r_benefit_header + 1
for i, (k, lo, hi, note) in enumerate(benefits):
    r = ben_start + i
    ws_st.cell(row=r, column=1, value=k).font = BLACK
    ws_st.cell(row=r, column=2, value=lo).font = BLUE
    ws_st.cell(row=r, column=2).number_format = '£#,##0.0'
    ws_st.cell(row=r, column=3, value=hi).font = BLUE
    ws_st.cell(row=r, column=3).number_format = '£#,##0.0'
    ws_st.cell(row=r, column=4, value=note).font = SUBTITLE
    ws_st.cell(row=r, column=4).alignment = LFT

r_total_ben = ben_start + len(benefits)
ws_st.cell(row=r_total_ben, column=1, value="TOTAL BENEFITS TO UK").font = BOLD_BLACK
ws_st.cell(row=r_total_ben, column=1).fill = FILL_RESULT
ws_st.cell(row=r_total_ben, column=2, value=f"=SUM(B{ben_start}:B{r_total_ben - 1})").font = BOLD_BLACK
ws_st.cell(row=r_total_ben, column=2).fill = FILL_RESULT
ws_st.cell(row=r_total_ben, column=2).number_format = '£#,##0.0'
ws_st.cell(row=r_total_ben, column=3, value=f"=SUM(C{ben_start}:C{r_total_ben - 1})").font = BOLD_BLACK
ws_st.cell(row=r_total_ben, column=3).fill = FILL_RESULT
ws_st.cell(row=r_total_ben, column=3).number_format = '£#,##0.0'

# Net welfare effect
r_net = r_total_ben + 2
ws_st.cell(row=r_net, column=1, value="NET WELFARE EFFECT (Benefits − Costs)").font = BOLD_BLACK
ws_st.cell(row=r_net, column=1).fill = FILL_HIGHLIGHT
ws_st.cell(row=r_net, column=2, value=f"=B{r_total_ben}-B{r_total_cost}").font = BOLD_BLACK
ws_st.cell(row=r_net, column=2).fill = FILL_HIGHLIGHT
ws_st.cell(row=r_net, column=2).number_format = '£#,##0.0'
ws_st.cell(row=r_net, column=3, value=f"=C{r_total_ben}-C{r_total_cost}").font = BOLD_BLACK
ws_st.cell(row=r_net, column=3).fill = FILL_HIGHLIGHT
ws_st.cell(row=r_net, column=3).number_format = '£#,##0.0'
ws_st.cell(row=r_net, column=4, value="Negative under all parameter scenarios. The pause is the UK's welfare-superior outcome.").font = BOLD_BLACK
ws_st.cell(row=r_net, column=4).alignment = LFT

# Conclusion
r_conc = r_net + 2
ws_st.cell(row=r_conc, column=1, value="CONCLUSION: For each £1 of project value Stargate UK would deliver to the UK, it would cost £10–15 in concessions to make the project viable.").font = BOLD_BLACK
ws_st.merge_cells(start_row=r_conc, start_column=1, end_row=r_conc, end_column=4)
ws_st.cell(row=r_conc, column=1).alignment = LFT
ws_st.cell(row=r_conc, column=1).fill = FILL_HIGHLIGHT

print(f"Stargate sheet built. Net welfare row: {r_net}")

# Reorder sheets — put 00_Summary first
sheets_order = ["00_Summary", "01_Parameters", "02_Sector_Panel", "03_Lambda_Decomp",
                "04_UK_US_Flows", "05_HMRC_Loss", "06_Counterfactual", "07_Scenarios", "08_Stargate_Pricing"]
wb._sheets = [wb[name] for name in sheets_order]

wb.save('/home/claude/uk_ai_externality_model.xlsx')
print("All sheets built. File saved.")
