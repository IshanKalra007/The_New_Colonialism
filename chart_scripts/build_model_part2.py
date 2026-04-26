"""
Part 2: Sector Panel — quarterly headcount, AI spend, displacement,
productivity, wage bill, tax loss for 8 UK sectors Q1 2023 to Q4 2027.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = load_workbook('/home/claude/uk_ai_externality_model.xlsx')

ARIAL = "Arial"
BLUE = Font(name=ARIAL, color="0000FF", size=10)
BLACK = Font(name=ARIAL, color="000000", size=10)
GREEN = Font(name=ARIAL, color="008000", size=10)
BOLD_BLACK = Font(name=ARIAL, color="000000", size=10, bold=True)
HEADER = Font(name=ARIAL, color="FFFFFF", size=11, bold=True)
TITLE = Font(name=ARIAL, color="000000", size=14, bold=True)
SUBTITLE = Font(name=ARIAL, color="555555", size=10, italic=True)
FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FILL_SUBHEAD = PatternFill("solid", start_color="D9E1F2")
FILL_HIGHLIGHT = PatternFill("solid", start_color="FFF2CC")
FILL_RESULT = PatternFill("solid", start_color="E2EFDA")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
FMT_GBP_M = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"_-;_-@_-'
FMT_PCT = '0.0%'
FMT_NUM = '#,##0'

ws = wb.create_sheet("02_Sector_Panel")

# ----------------------------------------------------------------------------
# Sector definitions
# Headcount Q1 2023 baseline + sector params
# ----------------------------------------------------------------------------
# Each row: (sector, headcount_Q1_23, alpha_Q1_23, alpha_Q1_26, beta_central, eta_central, avg_wage_£, sector_color)
sectors = [
    ("Big 4 prof svcs",    71000,  0.65, 0.92, 3.3,  0.85, 65000),
    ("UK banking",        315000,  0.45, 0.78, 6.7,  0.45, 55000),
    ("UK telco",          165000,  0.30, 0.62, 7.1,  0.30, 45000),
    ("UK retail",        1180000,  0.15, 0.38, 7.2,  0.15, 32000),
    ("UK creative/PSB",   220000,  0.40, 0.68, 10.0, 0.55, 42000),
    ("UK public sector", 5460000,  0.10, 0.32, 4.5,  0.25, 38000),
    ("UK manufacturing", 2640000,  0.18, 0.40, 5.5,  0.18, 36000),
    ("UK SME aggregate", 8200000,  0.23, 0.62, 6.0,  0.40, 33000),
]

# 20 quarterly periods Q1 2023 to Q4 2027
quarters = []
for y in [2023, 2024, 2025, 2026, 2027]:
    for q in [1, 2, 3, 4]:
        quarters.append(f"{y}Q{q}")

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 18
for i in range(len(quarters)):
    ws.column_dimensions[get_column_letter(3 + i)].width = 11

# Title
ws.cell(row=1, column=1, value="Sector Panel — Quarterly Time Series Q1 2023 to Q4 2027").font = TITLE
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)
ws.cell(row=2, column=1, value="Headcount, AI adoption, AI spend, AI-attributed displacement, productivity, wage loss, HMRC tax loss for 8 UK sectors. All £m.").font = SUBTITLE
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=22)

# Header row
header_cells = ["Sector", "Metric"] + quarters
for i, h in enumerate(header_cells):
    c = ws.cell(row=4, column=1 + i, value=h)
    c.font = HEADER
    c.fill = FILL_HEADER
    c.alignment = CTR
    c.border = BORDER_ALL

# ----------------------------------------------------------------------------
# Build panel — 8 sectors × 6 metric rows = 48 data rows
# Metrics: Headcount, α, AI spend (£m/quarter), Cumulative AI layoffs,
#          Wage bill loss (£m/qtr), HMRC labour tax loss (£m/qtr)
# ----------------------------------------------------------------------------
# Adoption pattern: linear interpolation Q1 23 → Q1 26 + Bass-fitted Q1 26 → Q4 27
def alpha_path(a0, a3, a_inf):
    """Generate 20-quarter adoption series from baseline, Q1 26 calibration, asymptote."""
    series = []
    # Q1 23 to Q1 26: 13 quarters, linear from a0 to a3
    for i in range(13):
        f = i / 12
        series.append(a0 + (a3 - a0) * f)
    # Q2 26 to Q4 27: 7 quarters, Bass continuation toward asymptote
    for i in range(7):
        prev = series[-1]
        # Bass-like: increment proportional to remaining capacity
        delta = 0.08 * (a_inf - prev)
        series.append(min(prev + delta, a_inf - 0.01))
    return series

def quarterly_layoffs(headcount, alpha_series, beta, eta, sigma_q, attribution=0.5):
    """Compute quarterly AI-attributed layoffs from spend × β × attribution.
    sigma_q is £/seat/quarter (£23 × 3 = £69)."""
    layoffs = []
    for i, a in enumerate(alpha_series):
        # AI spend this quarter (£m) = α × headcount × η × σ_quarterly / 1e6
        ai_spend_m = (a * headcount * eta * sigma_q) / 1e6
        # July 2026 kink: Q3 2026 onward = quarter 14 (0-indexed)
        if i >= 14:
            ai_spend_m *= 1.13
        # β is jobs cut per £1m AI spend per quarter
        # Apply attribution
        q_layoffs = ai_spend_m * beta * attribution
        layoffs.append(q_layoffs)
    return layoffs

def cumulative(series):
    out, total = [], 0
    for v in series:
        total += v
        out.append(total)
    return out

# Row tracker
r = 5

for s in sectors:
    name, hc0, a0, a3, beta, eta, wage = s

    # Use asymptote based on sector
    a_inf = 0.85 if hc0 < 1e6 else 0.65 if "SME" in name else 0.78

    alphas = alpha_path(a0, a3, a_inf)
    sigma_q = 23 * 3  # £69 per seat per quarter (enterprise)

    # AI spend per quarter (£m)
    spends = []
    for i, a in enumerate(alphas):
        ai_spend = (a * hc0 * eta * sigma_q) / 1e6
        if i >= 14:
            ai_spend *= 1.13
        spends.append(ai_spend)

    # Cumulative layoffs (50% attribution)
    quarterly_lo = quarterly_layoffs(hc0, alphas, beta, eta, sigma_q, attribution=0.5)
    cum_lo = cumulative(quarterly_lo)

    # Headcount = baseline minus cumulative AI layoffs (ignoring other change)
    hc_series = [hc0 - cl for cl in cum_lo]

    # Wage bill loss per quarter (£m) = quarterly layoffs × wage / 1e6 / 4 (only 1/4 year flow lost in incremental Q)
    # But we want cumulative wage bill loss too, treated as annualized loss running rate
    # Use quarterly_lo × annual wage as cumulative annual wage loss running rate
    wage_loss = [cl * wage / 1e6 for cl in cum_lo]

    # HMRC tax loss per quarter (£m) = wage_loss × τ_w (0.31)
    hmrc_loss = [w * 0.31 for w in wage_loss]

    # Write rows: Headcount, α, AI spend, Cum layoffs, Wage bill loss, HMRC loss
    sector_rows = [
        ("Headcount (FTE)",          hc_series, "0,0"),
        ("AI adoption rate (α)",      alphas,    "pct"),
        ("AI spend (£m / qtr)",       spends,    "gbp"),
        ("Cumulative AI layoffs",     cum_lo,    "0,0"),
        ("Annual wage bill lost (£m)", wage_loss, "gbp"),
        ("Annual HMRC tax loss (£m)",  hmrc_loss, "gbp"),
    ]

    # Sector name in column 1, merged
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    ws.cell(row=r, column=1).fill = FILL_SUBHEAD
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 5, end_column=1)
    ws.cell(row=r, column=1).alignment = CTR

    for j, (metric, vals, fmt) in enumerate(sector_rows):
        rr = r + j
        ws.cell(row=rr, column=2, value=metric).font = BLACK
        ws.cell(row=rr, column=2).alignment = LFT

        for i, v in enumerate(vals):
            c = ws.cell(row=rr, column=3 + i, value=v)
            c.font = BLUE if j == 0 else BLACK
            if fmt == "pct":
                c.number_format = FMT_PCT
            elif fmt == "gbp":
                c.number_format = FMT_GBP_M
            else:
                c.number_format = FMT_NUM
        # Highlight key metrics
        if metric == "Annual HMRC tax loss (£m)":
            for i in range(len(vals)):
                ws.cell(row=rr, column=3 + i).fill = FILL_RESULT

    r += 7  # 6 metrics + 1 spacer

# Add aggregate row
r_agg = r
ws.cell(row=r_agg, column=1, value="UK AGGREGATE").font = BOLD_BLACK
ws.cell(row=r_agg, column=1).fill = FILL_HIGHLIGHT
ws.merge_cells(start_row=r_agg, start_column=1, end_row=r_agg + 3, end_column=1)
ws.cell(row=r_agg, column=1).alignment = CTR

# Aggregate metrics
agg_metrics = [
    "Total AI spend (£m / qtr)",
    "Total cumulative AI layoffs",
    "Total annual wage bill lost (£m)",
    "Total annual HMRC tax loss (£m)",
]
# Source rows for sums (3rd, 4th, 5th, 6th metric in each sector block)
source_rows = [7, 8, 9, 10]  # AI spend, cum layoffs, wage loss, HMRC loss in row 5+
# Sector blocks start at row 5 then 12 then 19 ... step 7
sector_starts = [5 + 7 * k for k in range(len(sectors))]
# Within each block, AI spend is offset 2 (row+2), cum layoffs offset 3, wage loss offset 4, HMRC loss offset 5

for j, m in enumerate(agg_metrics):
    rr = r_agg + j
    ws.cell(row=rr, column=2, value=m).font = BOLD_BLACK
    ws.cell(row=rr, column=2).alignment = LFT
    ws.cell(row=rr, column=2).fill = FILL_HIGHLIGHT
    offset = j + 2  # AI spend = +2, cum layoffs = +3, wage loss = +4, HMRC loss = +5
    for i in range(20):
        col = 3 + i
        # SUM across sectors at the appropriate metric row
        cells = [f"{get_column_letter(col)}{ss + offset}" for ss in sector_starts]
        formula = "=" + "+".join(cells)
        c = ws.cell(row=rr, column=col, value=formula)
        c.font = BOLD_BLACK
        c.fill = FILL_HIGHLIGHT
        if "wage" in m.lower() or "spend" in m.lower() or "tax" in m.lower():
            c.number_format = FMT_GBP_M
        else:
            c.number_format = FMT_NUM

# Freeze panes
ws.freeze_panes = "C5"

wb.save('/home/claude/uk_ai_externality_model.xlsx')
print(f"Sector panel built. Aggregate rows starting at {r_agg}")
print(f"Sector starts: {sector_starts}")
