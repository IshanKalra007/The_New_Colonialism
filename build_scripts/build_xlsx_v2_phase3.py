"""
Phase 3 of build_xlsx_v2.py — Lambda_Decomp and Stargate_Counterfactual

Adds two analytical sheets:
  - Lambda_Decomp (5-channel computation of UK domestic capture; implied λ by year)
  - Stargate_Counterfactual (concession bundle vs project value; welfare margin)
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Style constants matching Phases 1-2
FONT_INPUT = Font(name='Calibri', size=10, color='0000FF')
FONT_FORMULA = Font(name='Calibri', size=10, color='000000')
FONT_LINK = Font(name='Calibri', size=10, color='008000')
FONT_NAMED = Font(name='Calibri', size=10, color='7030A0')
FONT_BOLD = Font(name='Calibri', size=10, bold=True, color='000000')
FONT_ITALIC = Font(name='Calibri', size=9, italic=True, color='666666')
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Calibri', size=16, bold=True, color='1F3A5F')
FONT_SECTION = Font(name='Calibri', size=11, bold=True, color='1F3A5F')

FILL_HEADER = PatternFill('solid', start_color='1F3A5F')
FILL_INPUT_BG = PatternFill('solid', start_color='FFF8DC')
FILL_OUTPUT_BG = PatternFill('solid', start_color='E8E8E8')
FILL_KEY_ASSUMPTION = PatternFill('solid', start_color='FFEB9C')
FILL_HEADLINE = PatternFill('solid', start_color='C6E0B4')
FILL_RECON_OK = PatternFill('solid', start_color='C6EFCE')
FILL_RECON_FAIL = PatternFill('solid', start_color='FFC7CE')
FILL_SECTION_BAND = PatternFill('solid', start_color='D9E1F2')

BORDER_TOTAL = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='double', color='000000')
)
BORDER_SECTION = Border(bottom=Side(style='medium', color='1F3A5F'))

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

FMT_GBP_BN = '"£"#,##0.0"bn";("£"#,##0.0"bn");"-"'
FMT_GBP_M = '"£"#,##0"m";("£"#,##0"m");"-"'
FMT_GBP_FULL = '"£"#,##0;("£"#,##0);"-"'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_DEC2 = '0.00'
FMT_DEC3 = '0.000'
FMT_INT = '#,##0'
FMT_YEAR = '0'


def section_band(ws, row, col_start, col_end, title):
    ws.cell(row=row, column=col_start, value=title).font = FONT_SECTION
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION_BAND
    ws.cell(row=row, column=col_start).border = BORDER_SECTION


PATH = '/mnt/user-data/outputs/replication/uk_ai_externality_model.xlsx'
wb = load_workbook(PATH)


# ============================================================
# SHEET: Lambda_Decomp
# ============================================================

ws = wb.create_sheet('Lambda_Decomp')
ws.sheet_properties.tabColor = '8FAADC'
ws.column_dimensions['A'].width = 38
for col_idx in range(2, 8):
    ws.column_dimensions[get_column_letter(col_idx)].width = 13

ws['A1'] = 'λ Decomposition: Five Channels of UK Domestic Capture'
ws['A1'].font = FONT_TITLE

ws['A2'] = 'Implied λ = 1 - (UK domestic capture / gross UK→US AI flow). Paper §5.2.'
ws['A2'].font = FONT_ITALIC
ws.merge_cells('A2:G2')

# Year header
years_lambda = [2024, 2025, 2026]
ws.cell(row=4, column=1, value='Channel').font = FONT_HEADER
ws.cell(row=4, column=1).fill = FILL_HEADER
for i, yr in enumerate(years_lambda):
    c = ws.cell(row=4, column=2+i, value=yr)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = CENTER
    c.number_format = FMT_YEAR
ws.cell(row=4, column=5, value='Source').font = FONT_HEADER
ws.cell(row=4, column=5).fill = FILL_HEADER

# Section: gross flow
section_band(ws, 5, 1, 7, 'GROSS UK→US AI FLOW')

ws.cell(row=6, column=1, value='Gross flow (£bn)').font = FONT_BOLD
# Pull from C1_Subscription_Flow row 14 columns C, D, E (2024, 2025, 2026)
ws.cell(row=6, column=2, value='=C1_Subscription_Flow!C14')
ws.cell(row=6, column=2).font = FONT_LINK
ws.cell(row=6, column=2).number_format = FMT_GBP_BN
ws.cell(row=6, column=3, value='=C1_Subscription_Flow!D14')
ws.cell(row=6, column=3).font = FONT_LINK
ws.cell(row=6, column=3).number_format = FMT_GBP_BN
ws.cell(row=6, column=4, value='=C1_Subscription_Flow!E14')
ws.cell(row=6, column=4).font = FONT_LINK
ws.cell(row=6, column=4).number_format = FMT_GBP_BN
for c in [2, 3, 4]:
    ws.cell(row=6, column=c).alignment = RIGHT
ws.cell(row=6, column=5, value='From C1_Subscription_Flow').font = FONT_ITALIC

# Section: five channels
section_band(ws, 8, 1, 7, 'FIVE CHANNELS OF UK DOMESTIC CAPTURE (£m)')

channels = [
    ('Channel 1: US AI firms\' UK staff wages',
     [560, 595, 631],
     'Microsoft UK (7,500 staff), Google UK (5,000), AWS UK (3,000), etc. Loaded comp £85k × AI share 33%.'),
    ('Channel 2: UK integration partner margins',
     [210, 315, 420],
     'Nscale, Capita, Computacenter, Softcat. UK partner AI revenue × ~6% margin.'),
    ('Channel 3: Transfer-priced UK corporation tax',
     [110, 165, 240],
     'Microsoft UK Ltd, Google UK Ltd, Amazon UK Services. CT on AI-attributable UK profit.'),
    ('Channel 4: UK university research partnerships',
     [115, 130, 145],
     'Microsoft Research Cambridge, DeepMind UK, Anthropic London, OpenAI London.'),
    ('Channel 5: UK shareholder returns + CGT',
     [55, 100, 156],
     'UK pension fund + ISA + SIPP US Big Tech holdings ~£175bn × dividend yield + CGT.'),
]

for i, (name, vals, src) in enumerate(channels):
    r = 9 + i
    ws.cell(row=r, column=1, value=name).font = FONT_BOLD
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=2+j, value=v)
        c.font = FONT_INPUT
        c.fill = FILL_INPUT_BG
        c.number_format = FMT_GBP_M
        c.alignment = RIGHT
    ws.cell(row=r, column=5, value=src).font = FONT_ITALIC
    ws.cell(row=r, column=5).alignment = WRAP

# Total capture in £m
ws.cell(row=14, column=1, value='Total UK domestic capture (£m)').font = FONT_BOLD
ws.cell(row=14, column=1).fill = FILL_HEADLINE
for j in range(3):
    col = get_column_letter(2+j)
    cell = ws.cell(row=14, column=2+j, value=f'=SUM({col}9:{col}13)')
    cell.font = FONT_FORMULA
    cell.number_format = FMT_GBP_M
    cell.fill = FILL_HEADLINE
    cell.alignment = RIGHT
    cell.border = BORDER_TOTAL

# Total in £bn
ws.cell(row=15, column=1, value='Total UK domestic capture (£bn)').font = FONT_BOLD
for j in range(3):
    col = get_column_letter(2+j)
    cell = ws.cell(row=15, column=2+j, value=f'={col}14/1000')
    cell.font = FONT_FORMULA
    cell.number_format = FMT_GBP_BN
    cell.alignment = RIGHT

# Implied λ
section_band(ws, 17, 1, 7, 'IMPLIED LAMBDA BY YEAR')

ws.cell(row=18, column=1, value='Implied λ = 1 - (capture / gross)').font = FONT_BOLD
ws.cell(row=18, column=1).fill = FILL_KEY_ASSUMPTION
for j in range(3):
    col = get_column_letter(2+j)
    cell = ws.cell(row=18, column=2+j, value=f'=1-{col}15/{col}6')
    cell.font = FONT_FORMULA
    cell.number_format = FMT_DEC3
    cell.fill = FILL_KEY_ASSUMPTION
    cell.alignment = RIGHT

# Paper claim row
ws.cell(row=19, column=1, value='Paper claim').font = FONT_BOLD
paper_lambdas = [0.76, 0.83, 0.87]
for j, v in enumerate(paper_lambdas):
    cell = ws.cell(row=19, column=2+j, value=v)
    cell.font = FONT_INPUT
    cell.number_format = FMT_DEC3
    cell.alignment = RIGHT

# Status
ws.cell(row=20, column=1, value='Status').font = FONT_BOLD
for j in range(3):
    col = get_column_letter(2+j)
    cell = ws.cell(row=20, column=2+j, value=f'=IF(ABS({col}18-{col}19)<=0.02,"OK","CHECK")')
    cell.font = FONT_BOLD
    cell.alignment = CENTER
    addr = f'{col}20'
    ws.conditional_formatting.add(addr,
        FormulaRule(formula=[f'${col}$20="OK"'], fill=FILL_RECON_OK))
    ws.conditional_formatting.add(addr,
        FormulaRule(formula=[f'${col}$20="CHECK"'], fill=FILL_RECON_FAIL))

# Central case used in paper
section_band(ws, 22, 1, 7, 'CENTRAL CASE USED IN MODEL')
ws.cell(row=23, column=1, value='λ central (named: lambda_central)').font = FONT_BOLD
ws.cell(row=23, column=2, value='=lambda_central').font = FONT_NAMED
ws.cell(row=23, column=2).number_format = FMT_DEC3
ws.cell(row=23, column=2).alignment = RIGHT
ws.cell(row=23, column=2).fill = FILL_HEADLINE

ws.cell(row=24, column=1, value='λ low').font = FONT_BOLD
ws.cell(row=24, column=2, value=0.78).font = FONT_INPUT
ws.cell(row=24, column=2).number_format = FMT_DEC3

ws.cell(row=25, column=1, value='λ high').font = FONT_BOLD
ws.cell(row=25, column=2, value=0.90).font = FONT_INPUT
ws.cell(row=25, column=2).number_format = FMT_DEC3

ws.freeze_panes = 'B5'


# ============================================================
# SHEET: Stargate_Counterfactual
# ============================================================

ws = wb.create_sheet('Stargate_Counterfactual')
ws.sheet_properties.tabColor = 'A9D08E'
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 60

ws['A1'] = 'Stargate UK Counterfactual: Pause vs Revival'
ws['A1'].font = FONT_TITLE

ws['A2'] = 'Concession bundle (cost) vs project value (benefit). The pause was welfare-superior. Paper §6.2.'
ws['A2'].font = FONT_ITALIC
ws.merge_cells('A2:E2')

# Header
for col, header in [('A', 'Component'), ('B', 'Low (£bn)'), ('C', 'High (£bn)'),
                    ('D', 'Central (£bn)'), ('E', 'Pricing rationale')]:
    c = ws[f'{col}4']
    c.value = header
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = CENTER

# CONCESSION BUNDLE
section_band(ws, 6, 1, 5, 'CONCESSION BUNDLE PV (cost to UK if Stargate revived)')

concessions = [
    ('Energy subsidy (matching UK to Texas rate)',
     2.0, 2.5,
     'Matching £180/MWh UK to £40/MWh Texas for 250-300 MW × 10 years.'),
    ('TDM exception (text-and-data mining)',
     15.0, 22.0,
     'Creative industries AI-vulnerable subsegment £40-60bn GVA × 1% perpetual licensing equiv, 4% disc.'),
    ('Weakened AISI pre-deployment testing',
     4.0, 6.0,
     'Reduced compliance costs and faster deployment timelines.'),
    ('Statutory liability shielding',
     5.0, 8.0,
     'Reduced UK litigation exposure across regulated sectors.'),
    ('Public-procurement preferences',
     2.0, 4.0,
     'Preferential UK government AI contract access (~£8-15bn visible spend).'),
    ('GDPR Article 22 softening',
     3.0, 5.0,
     'Reduced compliance costs across financial services, health, public sector.'),
]

start_row = 7
for i, (name, lo, hi, rat) in enumerate(concessions):
    r = start_row + i
    ws.cell(row=r, column=1, value=name).font = FONT_BOLD
    cl = ws.cell(row=r, column=2, value=lo)
    cl.font = FONT_INPUT
    cl.fill = FILL_INPUT_BG
    cl.number_format = FMT_GBP_BN
    cl.alignment = RIGHT
    ch = ws.cell(row=r, column=3, value=hi)
    ch.font = FONT_INPUT
    ch.fill = FILL_INPUT_BG
    ch.number_format = FMT_GBP_BN
    ch.alignment = RIGHT
    cd = ws.cell(row=r, column=4, value=f'=AVERAGE(B{r}:C{r})')
    cd.font = FONT_FORMULA
    cd.number_format = FMT_GBP_BN
    cd.alignment = RIGHT
    cd.fill = FILL_OUTPUT_BG
    ws.cell(row=r, column=5, value=rat).font = FONT_ITALIC
    ws.cell(row=r, column=5).alignment = WRAP

# Total concession
total_conc_row = start_row + len(concessions)  # row 13
ws.cell(row=total_conc_row, column=1, value='Total concession bundle').font = FONT_BOLD
ws.cell(row=total_conc_row, column=1).fill = FILL_HEADLINE
for col_letter in ['B', 'C', 'D']:
    col_idx = ord(col_letter) - ord('A') + 1
    cell = ws.cell(row=total_conc_row, column=col_idx,
                   value=f'=SUM({col_letter}{start_row}:{col_letter}{total_conc_row-1})')
    cell.font = FONT_FORMULA
    cell.number_format = FMT_GBP_BN
    cell.fill = FILL_HEADLINE
    cell.alignment = RIGHT
    cell.border = BORDER_TOTAL

# PROJECT VALUE
section_band(ws, total_conc_row + 2, 1, 5, 'PROJECT VALUE PV (benefit to UK if Stargate proceeds)')

benefits_start = total_conc_row + 3

benefits = [
    ('Construction-phase value-add',
     1.0, 1.5,
     '30-40% domestic retention on £3.2bn build over 18-24 months.'),
    ('Marginal λ reduction for OpenAI specifically',
     2.0, 3.0,
     '200-400 ongoing operational roles, ~0.5% reduction in OpenAI-specific UK λ.'),
    ('Regional employment multipliers',
     0.5, 1.0,
     '200-400 direct + 600 indirect jobs in North East England.'),
    ('Sovereign compute access',
     0.0, 0.5,
     'Marginal value of UK-resident vs US-resident compute under existing OpenAI MoU.'),
]

for i, (name, lo, hi, rat) in enumerate(benefits):
    r = benefits_start + i
    ws.cell(row=r, column=1, value=name).font = FONT_BOLD
    cl = ws.cell(row=r, column=2, value=lo)
    cl.font = FONT_INPUT
    cl.fill = FILL_INPUT_BG
    cl.number_format = FMT_GBP_BN
    cl.alignment = RIGHT
    ch = ws.cell(row=r, column=3, value=hi)
    ch.font = FONT_INPUT
    ch.fill = FILL_INPUT_BG
    ch.number_format = FMT_GBP_BN
    ch.alignment = RIGHT
    cd = ws.cell(row=r, column=4, value=f'=AVERAGE(B{r}:C{r})')
    cd.font = FONT_FORMULA
    cd.number_format = FMT_GBP_BN
    cd.alignment = RIGHT
    cd.fill = FILL_OUTPUT_BG
    ws.cell(row=r, column=5, value=rat).font = FONT_ITALIC

total_proj_row = benefits_start + len(benefits)
ws.cell(row=total_proj_row, column=1, value='Total project value').font = FONT_BOLD
ws.cell(row=total_proj_row, column=1).fill = FILL_HEADLINE
for col_letter in ['B', 'C', 'D']:
    col_idx = ord(col_letter) - ord('A') + 1
    cell = ws.cell(row=total_proj_row, column=col_idx,
                   value=f'=SUM({col_letter}{benefits_start}:{col_letter}{total_proj_row-1})')
    cell.font = FONT_FORMULA
    cell.number_format = FMT_GBP_BN
    cell.fill = FILL_HEADLINE
    cell.alignment = RIGHT
    cell.border = BORDER_TOTAL

# WELFARE MARGIN
section_band(ws, total_proj_row + 2, 1, 5, 'WELFARE MARGIN: PAUSE BEATS REVIVAL')

margin_row = total_proj_row + 3
ws.cell(row=margin_row, column=1, value='Welfare margin (cost - benefit)').font = FONT_BOLD
ws.cell(row=margin_row, column=1).fill = FILL_KEY_ASSUMPTION

# Low margin = low cost - high benefit
ws.cell(row=margin_row, column=2,
        value=f'=B{total_conc_row}-C{total_proj_row}').font = FONT_FORMULA
ws.cell(row=margin_row, column=2).number_format = FMT_GBP_BN
ws.cell(row=margin_row, column=2).alignment = RIGHT
ws.cell(row=margin_row, column=2).fill = FILL_KEY_ASSUMPTION

# High margin = high cost - low benefit
ws.cell(row=margin_row, column=3,
        value=f'=C{total_conc_row}-B{total_proj_row}').font = FONT_FORMULA
ws.cell(row=margin_row, column=3).number_format = FMT_GBP_BN
ws.cell(row=margin_row, column=3).alignment = RIGHT
ws.cell(row=margin_row, column=3).fill = FILL_KEY_ASSUMPTION

# Central margin
ws.cell(row=margin_row, column=4,
        value=f'=D{total_conc_row}-D{total_proj_row}').font = FONT_FORMULA
ws.cell(row=margin_row, column=4).number_format = FMT_GBP_BN
ws.cell(row=margin_row, column=4).alignment = RIGHT
ws.cell(row=margin_row, column=4).fill = FILL_KEY_ASSUMPTION

ws.cell(row=margin_row, column=5,
        value='Pause beats revival by this margin').font = FONT_ITALIC

# Concessions:project ratio
ws.cell(row=margin_row+2, column=1,
        value='For every £1 of project value...').font = FONT_BOLD
ws.cell(row=margin_row+2, column=2,
        value=f'=B{total_conc_row}/C{total_proj_row}').font = FONT_FORMULA
ws.cell(row=margin_row+2, column=2).number_format = FMT_DEC2
ws.cell(row=margin_row+2, column=3,
        value=f'=C{total_conc_row}/B{total_proj_row}').font = FONT_FORMULA
ws.cell(row=margin_row+2, column=3).number_format = FMT_DEC2
ws.cell(row=margin_row+2, column=5,
        value='...UK pays this much in concessions to make Stargate viable').font = FONT_ITALIC

# Reconciliation
section_band(ws, margin_row+4, 1, 5, 'RECONCILIATION TO PAPER')

ws.cell(row=margin_row+5, column=1, value='Concession bundle (paper §6.2)').font = FONT_BOLD
ws.cell(row=margin_row+5, column=2, value=31).font = FONT_INPUT
ws.cell(row=margin_row+5, column=2).number_format = FMT_GBP_BN
ws.cell(row=margin_row+5, column=3, value=47).font = FONT_INPUT
ws.cell(row=margin_row+5, column=3).number_format = FMT_GBP_BN

ws.cell(row=margin_row+6, column=1, value='Project value (paper §6.2)').font = FONT_BOLD
ws.cell(row=margin_row+6, column=2, value=4).font = FONT_INPUT
ws.cell(row=margin_row+6, column=2).number_format = FMT_GBP_BN
ws.cell(row=margin_row+6, column=3, value=5).font = FONT_INPUT
ws.cell(row=margin_row+6, column=3).number_format = FMT_GBP_BN

ws.cell(row=margin_row+7, column=1, value='Welfare margin (paper §6.2)').font = FONT_BOLD
ws.cell(row=margin_row+7, column=2, value=26).font = FONT_INPUT
ws.cell(row=margin_row+7, column=2).number_format = FMT_GBP_BN
ws.cell(row=margin_row+7, column=3, value=44).font = FONT_INPUT
ws.cell(row=margin_row+7, column=3).number_format = FMT_GBP_BN

ws.freeze_panes = 'A5'


# ============================================================
# UPDATE Audit_Trail with Phase 3 stats
# ============================================================

ws = wb['Audit_Trail']
ws['C23'] = 12  # Now 12 sheets
ws['C26'] = 3   # Phase 3
ws['B28'] = 'Phase 3 status: 12 sheets — foundations + 6 components + λ-decomp + Stargate'
ws['B29'] = 'Phase 4 will add Sensitivity tables (two-way) and tornado diagram.'


# ============================================================
# SAVE
# ============================================================

desired_order = [
    'Dashboard', 'Assumptions', 'Scenario_Engine', 'Audit_Trail',
    'C1_Subscription_Flow', 'C2_Cloud', 'C3_Productivity_Rent',
    'C4_Displaced_Wage', 'C5_HMRC', 'C6_Forgone_Frontier',
    'Lambda_Decomp', 'Stargate_Counterfactual'
]
wb._sheets = [wb[name] for name in desired_order]

wb.save(PATH)
print(f'Phase 3 saved: {PATH}')
print(f'Sheet count: {len(wb.sheetnames)}')
print(f'Sheets: {wb.sheetnames}')
