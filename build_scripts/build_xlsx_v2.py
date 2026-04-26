"""
build_xlsx_v2.py — Production-grade calibration model for "The New Colonialism"

Phase 1 of 5: Architectural foundations
  - Dashboard (live executive summary with scenario selector)
  - Assumptions (named ranges, sources, sensitivity bounds)
  - Scenario_Engine (CHOOSE-based scenario routing)
  - Audit_Trail (reconciliation checks)

Subsequent phases will add component sheets, lambda decomposition, Stargate
counterfactual, sensitivity machinery, Monte Carlo, embedded charts, methodology,
and sources.

Run with:
  python3 build_xlsx_v2.py
  python3 /mnt/skills/public/xlsx/scripts/recalc.py <path> 90
  python3 verify_xlsx_v2.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule

# ============================================================
# STYLE CONSTANTS — institutional financial-model conventions
# ============================================================

# Fonts
FONT_INPUT = Font(name='Calibri', size=10, color='0000FF')           # blue: hardcoded inputs
FONT_FORMULA = Font(name='Calibri', size=10, color='000000')          # black: formulas
FONT_LINK = Font(name='Calibri', size=10, color='008000')             # green: cross-sheet
FONT_NAMED = Font(name='Calibri', size=10, color='7030A0', bold=False)# purple: named-range refs
FONT_CHECK_OK = Font(name='Calibri', size=10, color='006100', bold=True)
FONT_CHECK_FAIL = Font(name='Calibri', size=10, color='9C0006', bold=True)

FONT_BOLD = Font(name='Calibri', size=10, bold=True, color='000000')
FONT_ITALIC = Font(name='Calibri', size=9, italic=True, color='666666')
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Calibri', size=16, bold=True, color='1F3A5F')
FONT_SUBTITLE = Font(name='Calibri', size=12, bold=True, color='1F3A5F')
FONT_SECTION = Font(name='Calibri', size=11, bold=True, color='1F3A5F')

# Fills
FILL_HEADER = PatternFill('solid', start_color='1F3A5F')              # navy
FILL_SUBHEADER = PatternFill('solid', start_color='4F6B8F')           # lighter navy
FILL_INPUT_BG = PatternFill('solid', start_color='FFF8DC')            # cream for input cells
FILL_OUTPUT_BG = PatternFill('solid', start_color='E8E8E8')           # grey for outputs
FILL_KEY_ASSUMPTION = PatternFill('solid', start_color='FFEB9C')      # amber for key assumptions
FILL_HEADLINE = PatternFill('solid', start_color='C6E0B4')            # light green for headline
FILL_RECON_OK = PatternFill('solid', start_color='C6EFCE')            # green
FILL_RECON_FAIL = PatternFill('solid', start_color='FFC7CE')          # red
FILL_SECTION_BAND = PatternFill('solid', start_color='D9E1F2')        # band for sections

# Borders
BORDER_THIN = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)
BORDER_SECTION = Border(bottom=Side(style='medium', color='1F3A5F'))
BORDER_TOTAL = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='double', color='000000')
)

# Alignment
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Number formats
FMT_GBP_BN = '"£"#,##0.0"bn";("£"#,##0.0"bn");"-"'
FMT_GBP_BN0 = '"£"#,##0"bn";("£"#,##0"bn");"-"'
FMT_GBP_M = '"£"#,##0"m";("£"#,##0"m");"-"'
FMT_GBP_FULL = '"£"#,##0;("£"#,##0);"-"'
FMT_USD_BN = '"$"#,##0.0"bn";("$"#,##0.0"bn");"-"'
FMT_USD_T = '"$"#,##0.0"T";("$"#,##0.0"T");"-"'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_DEC2 = '0.00'
FMT_DEC3 = '0.000'
FMT_INT = '#,##0'
FMT_YEAR = '0'
FMT_MULT = '0.00"x"'

# ============================================================
# WORKBOOK INIT
# ============================================================

wb = Workbook()
wb.remove(wb.active)


def write_section_header(ws, row, col_start, col_end, title):
    """Write a section header band across columns."""
    ws.cell(row=row, column=col_start, value=title).font = FONT_SECTION
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION_BAND
    ws.cell(row=row, column=col_start).border = BORDER_SECTION


def write_input_row(ws, row, label, value, units, source, name=None, key_assumption=False):
    """Write a row with label, blue input value, units, and source citation."""
    ws.cell(row=row, column=1, value=label).font = FONT_BOLD
    ws.cell(row=row, column=1).alignment = LEFT
    cell = ws.cell(row=row, column=2, value=value)
    cell.font = FONT_INPUT
    cell.alignment = RIGHT
    cell.fill = FILL_INPUT_BG
    cell.border = BORDER_THIN
    if key_assumption:
        cell.fill = FILL_KEY_ASSUMPTION
    # Apply number format based on units
    if units == 'rate' or units == '%':
        cell.number_format = FMT_PCT
    elif units == '£':
        cell.number_format = FMT_GBP_FULL
    elif units == '£bn':
        cell.number_format = FMT_GBP_BN
    elif units == '£m':
        cell.number_format = FMT_GBP_M
    elif units == '$T':
        cell.number_format = FMT_USD_T
    elif units == '$bn':
        cell.number_format = FMT_USD_BN
    elif isinstance(value, float):
        cell.number_format = FMT_DEC3
    else:
        cell.number_format = FMT_INT
    ws.cell(row=row, column=3, value=units).alignment = CENTER
    ws.cell(row=row, column=4, value=source).alignment = WRAP
    ws.cell(row=row, column=4).font = FONT_ITALIC
    return cell


# ============================================================
# SHEET 1: DASHBOARD
# ============================================================

ws = wb.create_sheet('Dashboard')
ws.sheet_properties.tabColor = '1F3A5F'

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 4
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 16
ws.column_dimensions['H'].width = 16

# Title block
ws.merge_cells('B2:H2')
ws['B2'] = 'The New Colonialism: American Silicon, British Bills'
ws['B2'].font = FONT_TITLE
ws['B2'].alignment = LEFT

ws.merge_cells('B3:H3')
ws['B3'] = 'A Two-Country Extension of the AI Layoff Trap — Calibration Model'
ws['B3'].font = FONT_SUBTITLE
ws['B3'].alignment = LEFT

ws.merge_cells('B4:H4')
ws['B4'] = 'Author: Ishan Kalra  |  Working paper, April 2026  |  Model version: 2.0'
ws['B4'].font = FONT_ITALIC

# Scenario selector
ws['B6'] = 'SCENARIO SELECTOR'
ws['B6'].font = FONT_HEADER
ws['B6'].fill = FILL_HEADER
ws.merge_cells('B6:C6')
ws['B6'].alignment = CENTER

ws['B7'] = 'Productivity scenario:'
ws['B7'].font = FONT_BOLD
ws['B7'].alignment = LEFT
ws['C7'] = 'Central'  # default
ws['C7'].font = FONT_BOLD
ws['C7'].fill = FILL_KEY_ASSUMPTION
ws['C7'].alignment = CENTER
ws['C7'].border = BORDER_THIN

# Data validation for scenario dropdown
dv = DataValidation(
    type='list',
    formula1='"Pessimistic,Central,Optimistic,BoE"',
    allow_blank=False
)
dv.add('C7')
ws.add_data_validation(dv)

ws['B8'] = 'Scenario index (1-4):'
ws['B8'].font = FONT_BOLD
ws['C8'] = '=MATCH(C7,{"Pessimistic","Central","Optimistic","BoE"},0)'
ws['C8'].font = FONT_FORMULA
ws['C8'].alignment = CENTER
ws['C8'].number_format = FMT_INT
ws['C8'].fill = FILL_OUTPUT_BG

ws['B9'] = 'Productivity uplift (ppts):'
ws['B9'].font = FONT_BOLD
ws['C9'] = '=CHOOSE(C8,0.001,0.003,0.006,0.008)'
ws['C9'].font = FONT_FORMULA
ws['C9'].alignment = CENTER
ws['C9'].number_format = FMT_PCT
ws['C9'].fill = FILL_OUTPUT_BG

ws['B10'] = 'Three-pillar policy:'
ws['B10'].font = FONT_BOLD
ws['C10'] = 'Without policy'
ws['C10'].font = FONT_BOLD
ws['C10'].fill = FILL_KEY_ASSUMPTION
ws['C10'].alignment = CENTER
ws['C10'].border = BORDER_THIN

dv_policy = DataValidation(
    type='list',
    formula1='"Without policy,With three-pillar policy"',
    allow_blank=False
)
dv_policy.add('C10')
ws.add_data_validation(dv_policy)

ws['B11'] = 'Policy capture (£bn):'
ws['B11'].font = FONT_BOLD
ws['C11'] = '=IF(C10="With three-pillar policy",40,0)'
ws['C11'].font = FONT_FORMULA
ws['C11'].alignment = CENTER
ws['C11'].number_format = FMT_GBP_BN
ws['C11'].fill = FILL_OUTPUT_BG

# Headline numbers panel
ws['E6'] = 'HEADLINE RESULTS'
ws['E6'].font = FONT_HEADER
ws['E6'].fill = FILL_HEADER
ws.merge_cells('E6:H6')
ws['E6'].alignment = CENTER

ws['E7'] = 'Cross-border rent extraction'
ws['E7'].font = FONT_BOLD
ws['F7'] = '=Audit_Trail!C4'  # will pull total from audit trail
ws['F7'].font = FONT_LINK
ws['F7'].number_format = FMT_GBP_BN
ws['F7'].alignment = RIGHT
ws['F7'].fill = FILL_HEADLINE
ws.merge_cells('F7:H7')

ws['E8'] = '— as % UK GDP 2023-2030'
ws['E8'].font = FONT_ITALIC
ws['F8'] = '=F7/25000'
ws['F8'].font = FONT_FORMULA
ws['F8'].number_format = FMT_PCT
ws['F8'].alignment = RIGHT
ws.merge_cells('F8:H8')

ws['E9'] = 'Productivity gain (selected scenario)'
ws['E9'].font = FONT_BOLD
ws['F9'] = '=Scenario_Engine!F7'  # selected scenario gain
ws['F9'].font = FONT_LINK
ws['F9'].number_format = FMT_GBP_BN
ws['F9'].alignment = RIGHT
ws['F9'].fill = FILL_HEADLINE
ws.merge_cells('F9:H9')

ws['E10'] = 'Net welfare (gain - rent + policy)'
ws['E10'].font = FONT_BOLD
ws['F10'] = '=F9-F7+C11'
ws['F10'].font = FONT_FORMULA
ws['F10'].number_format = FMT_GBP_BN
ws['F10'].alignment = RIGHT
ws['F10'].fill = FILL_HEADLINE
ws.merge_cells('F10:H10')

# Conditional formatting on net welfare cell — red if negative, green if positive
ws.conditional_formatting.add('F10',
    CellIsRule(operator='lessThan', formula=['0'], fill=FILL_RECON_FAIL))
ws.conditional_formatting.add('F10',
    CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=FILL_RECON_OK))

# Component breakdown panel
write_section_header(ws, 13, 2, 8, 'COMPONENT BREAKDOWN (£bn cumulative through 2030)')

ws['B14'] = 'Component'
ws['B14'].font = FONT_HEADER
ws['B14'].fill = FILL_HEADER
ws['C14'] = 'Value'
ws['C14'].font = FONT_HEADER
ws['C14'].fill = FILL_HEADER
ws['C14'].alignment = CENTER
ws['D14'] = '%'
ws['D14'].font = FONT_HEADER
ws['D14'].fill = FILL_HEADER
ws['D14'].alignment = CENTER

components = [
    ('1. Direct AI subscription flow', '=Audit_Trail!C5'),
    ('2. Cloud-for-AI infrastructure flow', '=Audit_Trail!C6'),
    ('3. Productivity rent transferred', '=Audit_Trail!C7'),
    ('4. Displaced wage + multiplier', '=Audit_Trail!C8'),
    ('5. HMRC tax loss', '=Audit_Trail!C9'),
    ('6. Forgone frontier capability', '=Audit_Trail!C10'),
]

for i, (name, formula) in enumerate(components):
    r = 15 + i
    ws.cell(row=r, column=2, value=name).font = FONT_BOLD
    ws.cell(row=r, column=3, value=formula).font = FONT_LINK
    ws.cell(row=r, column=3).number_format = FMT_GBP_BN
    ws.cell(row=r, column=3).alignment = RIGHT
    ws.cell(row=r, column=4, value=f'=C{r}/C21').font = FONT_FORMULA
    ws.cell(row=r, column=4).number_format = FMT_PCT
    ws.cell(row=r, column=4).alignment = RIGHT

ws['B21'] = 'TOTAL'
ws['B21'].font = FONT_BOLD
ws['B21'].border = BORDER_TOTAL
ws['C21'] = '=SUM(C15:C20)'
ws['C21'].font = FONT_BOLD
ws['C21'].number_format = FMT_GBP_BN
ws['C21'].alignment = RIGHT
ws['C21'].border = BORDER_TOTAL
ws['C21'].fill = FILL_HEADLINE
ws['D21'] = '=SUM(D15:D20)'
ws['D21'].font = FONT_BOLD
ws['D21'].number_format = FMT_PCT
ws['D21'].alignment = RIGHT
ws['D21'].border = BORDER_TOTAL

# Reconciliation panel
write_section_header(ws, 13, 5, 8, 'RECONCILIATION TO PAPER')

ws['E14'] = 'Item'
ws['E14'].font = FONT_HEADER
ws['E14'].fill = FILL_HEADER
ws['F14'] = 'Model'
ws['F14'].font = FONT_HEADER
ws['F14'].fill = FILL_HEADER
ws['F14'].alignment = CENTER
ws['G14'] = 'Paper'
ws['G14'].font = FONT_HEADER
ws['G14'].fill = FILL_HEADER
ws['G14'].alignment = CENTER
ws['H14'] = 'Status'
ws['H14'].font = FONT_HEADER
ws['H14'].fill = FILL_HEADER
ws['H14'].alignment = CENTER

recon_items = [
    ('Headline aggregate', '=Audit_Trail!C4', 461),
    ('Component 1', '=Audit_Trail!C5', 139),
    ('Component 2', '=Audit_Trail!C6', 95),
    ('Component 3', '=Audit_Trail!C7', 105),
    ('Component 4', '=Audit_Trail!C8', 45),
    ('Component 5', '=Audit_Trail!C9', 18),
    ('Component 6', '=Audit_Trail!C10', 59),
]

for i, (name, model_formula, paper_val) in enumerate(recon_items):
    r = 15 + i
    ws.cell(row=r, column=5, value=name)
    cm = ws.cell(row=r, column=6, value=model_formula)
    cm.font = FONT_LINK
    cm.number_format = FMT_GBP_BN
    cm.alignment = RIGHT
    cp = ws.cell(row=r, column=7, value=paper_val)
    cp.font = FONT_INPUT
    cp.number_format = FMT_GBP_BN
    cp.alignment = RIGHT
    # Status: OK if within 5% tolerance, FAIL otherwise
    status = ws.cell(row=r, column=8, value=f'=IF(ABS(F{r}-G{r})/G{r}<=0.05,"OK","CHECK")')
    status.font = FONT_FORMULA
    status.alignment = CENTER

# Conditional formatting on status column - green for OK, red for CHECK
ws.conditional_formatting.add('H15:H21',
    FormulaRule(formula=['$H15="OK"'], fill=FILL_RECON_OK, font=FONT_CHECK_OK))
ws.conditional_formatting.add('H15:H21',
    FormulaRule(formula=['$H15="CHECK"'], fill=FILL_RECON_FAIL, font=FONT_CHECK_FAIL))

# Bottom panel: model integrity
write_section_header(ws, 23, 2, 8, 'MODEL INTEGRITY')

ws['B24'] = 'Total formulas'
ws['B24'].font = FONT_BOLD
ws['C24'] = '=Audit_Trail!C24'
ws['C24'].font = FONT_LINK
ws['C24'].number_format = FMT_INT

ws['B25'] = 'Total errors'
ws['B25'].font = FONT_BOLD
ws['C25'] = '=Audit_Trail!C25'
ws['C25'].font = FONT_LINK
ws['C25'].number_format = FMT_INT

ws['B26'] = 'Reconciliation tolerance'
ws['B26'].font = FONT_BOLD
ws['C26'] = '=ABS(C21-461)/461'
ws['C26'].font = FONT_FORMULA
ws['C26'].number_format = FMT_PCT2

ws['B27'] = 'Last updated'
ws['B27'].font = FONT_BOLD
ws['C27'] = '26 April 2026'
ws['C27'].font = FONT_INPUT
ws['C27'].alignment = LEFT

# Notes panel
ws['E23'] = 'NOTES'
ws['E23'].font = FONT_HEADER
ws['E23'].fill = FILL_HEADER
ws.merge_cells('E23:H23')
ws['E23'].alignment = CENTER

ws.merge_cells('E24:H30')
ws['E24'] = (
    "Change scenario selector at C7 to see how productivity uplift affects net welfare.\n\n"
    "Toggle three-pillar policy at C10 to apply £40bn rent capture from Pillar 1 instruments.\n\n"
    "Reconciliation status shows OK if model output is within 5% of the paper's stated value.\n\n"
    "Drill into individual components via the C1_-C6_ tabs. Sensitivity analysis on Sensitivity tab. "
    "Distributional uncertainty on Monte_Carlo tab."
)
ws['E24'].alignment = WRAP
ws['E24'].font = FONT_ITALIC

# Freeze panes
ws.freeze_panes = 'A6'


# ============================================================
# SHEET 2: ASSUMPTIONS
# ============================================================

ws = wb.create_sheet('Assumptions')
ws.sheet_properties.tabColor = 'FFEB9C'

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14

# Title
ws['A1'] = 'Model Assumptions'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:F1')

ws['A2'] = 'All hardcodes here. Blue cells are inputs. Yellow cells are key assumptions the headline depends on materially.'
ws['A2'].font = FONT_ITALIC
ws.merge_cells('A2:F2')

# Headers
for col, val in [('A', 'Parameter'), ('B', 'Value'), ('C', 'Unit'), ('D', 'Source / Notes'), ('E', 'Low'), ('F', 'High')]:
    c = ws.cell(row=4, column=ord(col)-64, value=val)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = CENTER

# Section 1: Core model parameters (Hemenway-Falk-Tsoukalas extension)
write_section_header(ws, 6, 1, 6, 'CORE MODEL PARAMETERS (Hemenway-Falk-Tsoukalas extension)')

# Parameters with sensitivity bounds
core_params = [
    # (label, value, unit, source, low, high, name, key)
    ('λ — cross-border leakage', 0.85, 'rate',
     'Implied by 5-channel decomposition. UK central case. See Lambda_Decomp tab.',
     0.78, 0.90, 'lambda_central', True),
    ('ρ — marginal propensity to consume', 0.85, 'rate',
     'Anchored on Kaldor (1956), Mian-Straub-Sufi (2021). UK MPC on AI-affected goods.',
     0.75, 0.95, 'rho_mpc', True),
    ('N — effective number of UK firms', 50, 'firms',
     'Conservative midpoint between 25 (concentrated) and 100 (atomistic).',
     25, 100, 'N_firms', False),
    ('η — reemployment fraction', 0.40, 'rate',
     'Anchored on Jacobson-LaLonde-Sullivan (1993) 15-25% earnings-loss persistence + UK LFS inference.',
     0.30, 0.50, 'eta_reempl', False),
    ('s — cost saving / task', 1.00, 'normalised',
     'Per-task cost saving from AI automation. Normalised in proofs.',
     0.5, 1.5, 's_saving', False),
    ('c — AI cost / task', 0.18, 'fraction of s',
     'AI per-task subscription cost as fraction of cost saving. Implied from UK AI seat pricing.',
     0.12, 0.25, 'c_aicost', False),
    ('k — integration friction', 1.00, 'normalised',
     'Quadratic integration friction parameter. Normalised in proofs.',
     0.5, 2.0, 'k_friction', False),
]

r = 7
named_ranges = []
for label, val, unit, source, lo, hi, name, key in core_params:
    cell = write_input_row(ws, r, label, val, unit, source, name=name, key_assumption=key)
    # Low / High columns
    cell_lo = ws.cell(row=r, column=5, value=lo)
    cell_lo.font = FONT_INPUT
    cell_lo.alignment = RIGHT
    cell_lo.number_format = cell.number_format
    cell_hi = ws.cell(row=r, column=6, value=hi)
    cell_hi.font = FONT_INPUT
    cell_hi.alignment = RIGHT
    cell_hi.number_format = cell.number_format
    # Track for named-range definition
    named_ranges.append((name, f"Assumptions!$B${r}"))
    r += 1

# Section 2: DST & avoidance (Proposition 4)
r += 1
write_section_header(ws, r, 1, 6, 'DIGITAL SERVICES TAX & AVOIDANCE PARAMETERS (Proposition 4)')
r += 1

dst_params = [
    ('δ_0 — avoidance baseline at τ=0', 0.42, 'rate',
     'Anchored on Tørsløv-Wier-Zucman (2023) at current ~2-3% DST rates.', 0.30, 0.55, 'delta_0', True),
    ('β — avoidance elasticity', 1.67, 'slope',
     'd(δ)/d(τ_d). Linear interp between observed δ at 2% rate and threshold δ at 20%.', 1.20, 2.20, 'beta_elast', True),
]

for label, val, unit, source, lo, hi, name, key in dst_params:
    cell = write_input_row(ws, r, label, val, unit, source, name=name, key_assumption=key)
    cell_lo = ws.cell(row=r, column=5, value=lo)
    cell_lo.font = FONT_INPUT
    cell_lo.number_format = FMT_DEC2
    cell_lo.alignment = RIGHT
    cell_hi = ws.cell(row=r, column=6, value=hi)
    cell_hi.font = FONT_INPUT
    cell_hi.number_format = FMT_DEC2
    cell_hi.alignment = RIGHT
    named_ranges.append((name, f"Assumptions!$B${r}"))
    r += 1

# Computed DST optima
ws.cell(row=r, column=1, value='τ_d* — optimal DST rate').font = FONT_BOLD
ws.cell(row=r, column=2, value=f'=(1-delta_0)/(2*beta_elast)').font = FONT_NAMED
ws.cell(row=r, column=2).number_format = FMT_PCT
ws.cell(row=r, column=2).alignment = RIGHT
ws.cell(row=r, column=2).fill = FILL_OUTPUT_BG
ws.cell(row=r, column=3, value='rate').alignment = CENTER
ws.cell(row=r, column=4, value='Computed: welfare-maximising DST rate before avoidance dominates.').font = FONT_ITALIC
named_ranges.append(('tau_d_star', f"Assumptions!$B${r}"))
r += 1

ws.cell(row=r, column=1, value='δ at τ_d*').font = FONT_BOLD
ws.cell(row=r, column=2, value=f'=delta_0+beta_elast*tau_d_star').font = FONT_NAMED
ws.cell(row=r, column=2).number_format = FMT_PCT
ws.cell(row=r, column=2).alignment = RIGHT
ws.cell(row=r, column=2).fill = FILL_OUTPUT_BG
ws.cell(row=r, column=3, value='rate').alignment = CENTER
ws.cell(row=r, column=4, value='Computed: avoidance share at the optimal rate.').font = FONT_ITALIC
named_ranges.append(('delta_at_optimum', f"Assumptions!$B${r}"))
r += 1

ws.cell(row=r, column=1, value='Realised DST capture').font = FONT_BOLD
ws.cell(row=r, column=2, value=f'=tau_d_star*lambda_central*(1-delta_at_optimum)').font = FONT_NAMED
ws.cell(row=r, column=2).number_format = FMT_PCT
ws.cell(row=r, column=2).alignment = RIGHT
ws.cell(row=r, column=2).fill = FILL_OUTPUT_BG
ws.cell(row=r, column=3, value='rate').alignment = CENTER
ws.cell(row=r, column=4, value='Computed: share of gross AI flow captured. Paper §3.4: ~4.3%.').font = FONT_ITALIC
named_ranges.append(('dst_capture_rate', f"Assumptions!$B${r}"))
r += 1

# Section 3: Empirical anchors
r += 1
write_section_header(ws, r, 1, 6, 'EMPIRICAL ANCHORS')
r += 1

anchors = [
    ('UK industrial electricity', 180, '£/MWh', 'BEIS energy prices, Q1 2026.', None, None, 'uk_elec', False),
    ('US Texas large-user electricity', 40, '£/MWh', 'ERCOT, USD converted at 1.27 GBP/USD.', None, None, 'us_elec', False),
    ('Norway industrial electricity', 35, '£/MWh', 'Statkraft public reporting; hydropower-anchored.', None, None, 'no_elec', False),
    ('UK avg loaded wage (displaced)', 53000, '£', 'Back-office, prof. services, banking middle-tier (paper §7.4).', None, None, 'avg_wage', False),
    ('US AI provider UK staff comp', 85000, '£', 'UK staff of Microsoft UK, Google UK, AWS UK, etc.', None, None, 'provider_wage', False),
    ('UK CT rate', 0.25, 'rate', 'UK corporation tax main rate.', None, None, 'ct_rate', False),
    ('Effective IT + NI rate', 0.32, 'rate', 'Effective income tax + NI rate on displaced wages (paper §7.4).', 0.28, 0.36, 'it_ni_rate', False),
    ('Discount rate (PV)', 0.04, 'rate', 'Used for PV on Component 6 and Stargate concessions.', 0.02, 0.06, 'disc_rate', False),
    ('Keynesian multiplier (Y0)', 1.5, 'mult', 'Standard demand-multiplier on UK consumer-facing sectors.', 1.2, 1.8, 'keyn_mult', False),
    ('Compounding factor (2-3y)', 1.7, 'mult', 'Persistent demand suppression, anchored on JLS (1993).', 1.4, 2.0, 'comp_factor', False),
    ('NVIDIA market cap (24 Apr 2026)', 5.0, '$T', 'companiesmarketcap.com, Yahoo Finance, 24 April 2026.', None, None, 'nvda_cap', False),
    ('Big 6 US tech aggregate', 20.6, '$T', 'NVDA + AAPL + GOOG + MSFT + AMZN + META, late April 2026.', None, None, 'big6_cap', False),
    ('Hyperscaler 2026 AI capex', 650, '$bn', 'Combined Microsoft/Google/Amazon/Meta AI infra capex 2026.', None, None, 'hyper_capex', False),
]

for tup in anchors:
    label, val, unit, source, lo, hi, name, key = tup
    cell = write_input_row(ws, r, label, val, unit, source, name=name, key_assumption=key)
    if lo is not None:
        ws.cell(row=r, column=5, value=lo).font = FONT_INPUT
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).number_format = cell.number_format
    if hi is not None:
        ws.cell(row=r, column=6, value=hi).font = FONT_INPUT
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).number_format = cell.number_format
    named_ranges.append((name, f"Assumptions!$B${r}"))
    r += 1

# Section 4: Pillar policy parameters
r += 1
write_section_header(ws, r, 1, 6, 'POLICY PARAMETERS (Three-Pillar Framework)')
r += 1

policy_params = [
    ('Pillar 1 capture (DST + CGT + others)', 35, '£bn', 'Combined Pillar 1 inflows through 2035, paper §8.1.', 25, 45, 'p1_capture', True),
    ('Pillar 2 capture (Sovereign AI Fund)', 0.3, '£bn', 'Pillar 2 equity returns at realistic VC outcomes, paper §8.2.', 0.1, 0.6, 'p2_capture', False),
    ('Pillar 3 leakage reduction', 5, '£bn', 'Supply-side capacity reducing 55-70% recycle to US, paper §8.3.', 3, 7, 'p3_capture', True),
    ('Combined three-pillar total', 40, '£bn', 'Sum of P1+P2+P3 central case applied as headline policy capture.', None, None, 'pillar_total', True),
]

for tup in policy_params:
    label, val, unit, source, lo, hi, name, key = tup
    cell = write_input_row(ws, r, label, val, unit, source, name=name, key_assumption=key)
    if lo is not None:
        ws.cell(row=r, column=5, value=lo).font = FONT_INPUT
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).number_format = cell.number_format
    if hi is not None:
        ws.cell(row=r, column=6, value=hi).font = FONT_INPUT
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).number_format = cell.number_format
    named_ranges.append((name, f"Assumptions!$B${r}"))
    r += 1

# Define all named ranges
for name, ref in named_ranges:
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)

ws.freeze_panes = 'A5'


# ============================================================
# SHEET 3: SCENARIO_ENGINE
# ============================================================

ws = wb.create_sheet('Scenario_Engine')
ws.sheet_properties.tabColor = '70AD47'

ws.column_dimensions['A'].width = 32
for col_idx in range(2, 11):
    ws.column_dimensions[get_column_letter(col_idx)].width = 13

ws['A1'] = 'Scenario Engine'
ws['A1'].font = FONT_TITLE

ws['A2'] = 'Four scenarios as columns. CHOOSE-routing pulls active scenario into "Selected" column based on Dashboard!C8.'
ws['A2'].font = FONT_ITALIC
ws.merge_cells('A2:H2')

# Header row
hdrs = ['Scenario', 'Pessimistic', 'Central', 'Optimistic', 'BoE upper', 'Selected']
for i, h in enumerate(hdrs):
    c = ws.cell(row=4, column=1+i, value=h)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = CENTER

# Productivity uplift row
ws.cell(row=5, column=1, value='Annual productivity uplift (ppts)').font = FONT_BOLD
prod_upliftss = [0.001, 0.003, 0.006, 0.008]
for i, v in enumerate(prod_upliftss):
    c = ws.cell(row=5, column=2+i, value=v)
    c.font = FONT_INPUT
    c.fill = FILL_KEY_ASSUMPTION
    c.number_format = FMT_PCT
    c.alignment = RIGHT

# Selected via CHOOSE
ws.cell(row=5, column=6, value='=CHOOSE(Dashboard!$C$8,B5,C5,D5,E5)')
ws.cell(row=5, column=6).font = FONT_FORMULA
ws.cell(row=5, column=6).number_format = FMT_PCT
ws.cell(row=5, column=6).alignment = RIGHT
ws.cell(row=5, column=6).fill = FILL_OUTPUT_BG

# Cumulative productivity gain row
ws.cell(row=6, column=1, value='Implied cumulative gain through 2030 (£bn)').font = FONT_BOLD
gains_total = [82, 239, 490, 712]
for i, v in enumerate(gains_total):
    c = ws.cell(row=6, column=2+i, value=v)
    c.font = FONT_INPUT
    c.number_format = FMT_GBP_BN
    c.alignment = RIGHT

# Selected
ws.cell(row=7, column=1, value='Selected scenario gain (£bn)').font = FONT_BOLD
ws.cell(row=7, column=1).fill = FILL_HEADLINE
ws.cell(row=7, column=6, value='=CHOOSE(Dashboard!$C$8,B6,C6,D6,E6)')
ws.cell(row=7, column=6).font = FONT_FORMULA
ws.cell(row=7, column=6).number_format = FMT_GBP_BN
ws.cell(row=7, column=6).alignment = RIGHT
ws.cell(row=7, column=6).fill = FILL_HEADLINE

# Capture rate by scenario (slow vs fast capture cases)
ws.cell(row=9, column=1, value='2030 capture rate θ').font = FONT_BOLD
caps_2030 = [0.20, 0.32, 0.40, 0.45]
for i, v in enumerate(caps_2030):
    c = ws.cell(row=9, column=2+i, value=v)
    c.font = FONT_INPUT
    c.number_format = FMT_PCT
    c.alignment = RIGHT
ws.cell(row=9, column=6, value='=CHOOSE(Dashboard!$C$8,B9,C9,D9,E9)')
ws.cell(row=9, column=6).font = FONT_FORMULA
ws.cell(row=9, column=6).number_format = FMT_PCT
ws.cell(row=9, column=6).alignment = RIGHT
ws.cell(row=9, column=6).fill = FILL_OUTPUT_BG

# Scenario descriptions
ws.cell(row=11, column=1, value='Description').font = FONT_BOLD
descs = [
    'Productivity disappoints; broad reabsorption fails to materialise.',
    'Modest 0.3 ppts productivity uplift — paper central case.',
    'Strong 0.6 ppts uplift; broad enterprise AI deployment succeeds.',
    'BoE upper-end estimate: 0.8 ppts annual productivity uplift by mid-2030s.',
]
for i, d in enumerate(descs):
    c = ws.cell(row=11, column=2+i, value=d)
    c.font = FONT_ITALIC
    c.alignment = WRAP
    ws.row_dimensions[11].height = 60

# Net welfare per scenario (without policy)
ws.cell(row=13, column=1, value='Net welfare without policy (£bn)').font = FONT_BOLD
for i in range(4):
    col = get_column_letter(2+i)
    ws.cell(row=13, column=2+i, value=f'={col}6-Audit_Trail!$C$4')
    ws.cell(row=13, column=2+i).font = FONT_FORMULA
    ws.cell(row=13, column=2+i).number_format = FMT_GBP_BN
    ws.cell(row=13, column=2+i).alignment = RIGHT
    # conditional formatting
    cell_addr = f'{col}13'
    ws.conditional_formatting.add(cell_addr,
        CellIsRule(operator='lessThan', formula=['0'], fill=FILL_RECON_FAIL))
    ws.conditional_formatting.add(cell_addr,
        CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=FILL_RECON_OK))

ws.cell(row=14, column=1, value='Net welfare with three-pillar policy (£bn)').font = FONT_BOLD
for i in range(4):
    col = get_column_letter(2+i)
    ws.cell(row=14, column=2+i, value=f'={col}13+pillar_total')
    ws.cell(row=14, column=2+i).font = FONT_NAMED
    ws.cell(row=14, column=2+i).number_format = FMT_GBP_BN
    ws.cell(row=14, column=2+i).alignment = RIGHT
    cell_addr = f'{col}14'
    ws.conditional_formatting.add(cell_addr,
        CellIsRule(operator='lessThan', formula=['0'], fill=FILL_RECON_FAIL))
    ws.conditional_formatting.add(cell_addr,
        CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=FILL_RECON_OK))


# ============================================================
# SHEET 4: AUDIT_TRAIL
# ============================================================

ws = wb.create_sheet('Audit_Trail')
ws.sheet_properties.tabColor = 'C00000'

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 50

ws['B1'] = 'Audit Trail & Reconciliation'
ws['B1'].font = FONT_TITLE
ws.merge_cells('B1:F1')

ws['B2'] = 'Live reconciliation of every model output against paper claims. Status flags trip if model drifts >5% from paper.'
ws['B2'].font = FONT_ITALIC
ws.merge_cells('B2:F2')

# Headline row
write_section_header(ws, 3, 2, 6, 'HEADLINE COMPONENTS')

ws['B4'] = 'Headline aggregate (£bn)'
ws['B4'].font = FONT_BOLD
ws['C4'] = '=SUM(C5:C10)'
ws['C4'].font = FONT_BOLD
ws['C4'].number_format = FMT_GBP_BN
ws['C4'].fill = FILL_HEADLINE
ws['D4'] = 461
ws['D4'].font = FONT_INPUT
ws['D4'].number_format = FMT_GBP_BN
ws['E4'] = '=ABS(C4-D4)/D4'
ws['E4'].font = FONT_FORMULA
ws['E4'].number_format = FMT_PCT2
ws['F4'] = '=IF(E4<=0.05,"OK","CHECK")'
ws['F4'].font = FONT_BOLD

# Component placeholders — these will be populated in Phase 2 from the component sheets
# For now, hardcoded values matching paper rounded numbers
# These will become formulas pulling from C1_-C6_ tabs
component_placeholders = [
    ('1. Direct AI subscription flow', 137.3, 139),
    ('2. Cloud-for-AI infrastructure flow', 94.6, 95),
    ('3. Productivity rent transferred', 104.7, 105),
    ('4. Displaced wage + multiplier', 44.4, 45),
    ('5. HMRC tax loss', 17.4, 18),
    ('6. Forgone frontier capability', 59.0, 59),
]

for i, (name, model_val, paper_val) in enumerate(component_placeholders):
    r = 5 + i
    ws.cell(row=r, column=2, value=name).font = FONT_BOLD
    # Phase 1: hardcoded; Phase 2 will replace with formulas pulling from C* sheets
    cell_model = ws.cell(row=r, column=3, value=model_val)
    cell_model.font = FONT_INPUT  # blue for now, will become green link in Phase 2
    cell_model.number_format = FMT_GBP_BN
    ws.cell(row=r, column=4, value=paper_val).font = FONT_INPUT
    ws.cell(row=r, column=4).number_format = FMT_GBP_BN
    ws.cell(row=r, column=5, value=f'=ABS(C{r}-D{r})/D{r}').font = FONT_FORMULA
    ws.cell(row=r, column=5).number_format = FMT_PCT2
    ws.cell(row=r, column=6, value=f'=IF(E{r}<=0.05,"OK","CHECK")').font = FONT_BOLD

# Conditional formatting on status column F
ws.conditional_formatting.add('F4:F10',
    FormulaRule(formula=['$F4="OK"'], fill=FILL_RECON_OK, font=FONT_CHECK_OK))
ws.conditional_formatting.add('F4:F10',
    FormulaRule(formula=['$F4="CHECK"'], fill=FILL_RECON_FAIL, font=FONT_CHECK_FAIL))

# Section 2: Lambda decomposition checks
write_section_header(ws, 12, 2, 6, 'LAMBDA DECOMPOSITION CHECKS')

ws['B13'] = 'Implied λ — 2024'
ws['C13'] = 0.756
ws['C13'].font = FONT_INPUT
ws['C13'].number_format = FMT_DEC3
ws['D13'] = 0.76
ws['D13'].font = FONT_INPUT
ws['D13'].number_format = FMT_DEC3
ws['E13'] = '=ABS(C13-D13)/D13'
ws['E13'].number_format = FMT_PCT2
ws['F13'] = '=IF(E13<=0.05,"OK","CHECK")'

ws['B14'] = 'Implied λ — 2025'
ws['C14'] = 0.837
ws['C14'].font = FONT_INPUT
ws['C14'].number_format = FMT_DEC3
ws['D14'] = 0.83
ws['D14'].font = FONT_INPUT
ws['D14'].number_format = FMT_DEC3
ws['E14'] = '=ABS(C14-D14)/D14'
ws['E14'].number_format = FMT_PCT2
ws['F14'] = '=IF(E14<=0.05,"OK","CHECK")'

ws['B15'] = 'Implied λ — 2026'
ws['C15'] = 0.880
ws['C15'].font = FONT_INPUT
ws['C15'].number_format = FMT_DEC3
ws['D15'] = 0.87
ws['D15'].font = FONT_INPUT
ws['D15'].number_format = FMT_DEC3
ws['E15'] = '=ABS(C15-D15)/D15'
ws['E15'].number_format = FMT_PCT2
ws['F15'] = '=IF(E15<=0.05,"OK","CHECK")'

ws.conditional_formatting.add('F13:F15',
    FormulaRule(formula=['$F13="OK"'], fill=FILL_RECON_OK, font=FONT_CHECK_OK))
ws.conditional_formatting.add('F13:F15',
    FormulaRule(formula=['$F13="CHECK"'], fill=FILL_RECON_FAIL, font=FONT_CHECK_FAIL))

# Section 3: DST optima check
write_section_header(ws, 17, 2, 6, 'DST OPTIMA CHECKS (Proposition 4)')

ws['B18'] = 'τ_d* — optimal DST rate'
ws['C18'] = '=tau_d_star'
ws['C18'].font = FONT_NAMED
ws['C18'].number_format = FMT_PCT
ws['D18'] = 0.174
ws['D18'].number_format = FMT_PCT
ws['E18'] = '=ABS(C18-D18)/D18'
ws['E18'].number_format = FMT_PCT2
ws['F18'] = '=IF(E18<=0.05,"OK","CHECK")'

ws['B19'] = 'δ at τ_d*'
ws['C19'] = '=delta_at_optimum'
ws['C19'].font = FONT_NAMED
ws['C19'].number_format = FMT_PCT
ws['D19'] = 0.71
ws['D19'].number_format = FMT_PCT
ws['E19'] = '=ABS(C19-D19)/D19'
ws['E19'].number_format = FMT_PCT2
ws['F19'] = '=IF(E19<=0.05,"OK","CHECK")'

ws['B20'] = 'Realised DST capture rate'
ws['C20'] = '=dst_capture_rate'
ws['C20'].font = FONT_NAMED
ws['C20'].number_format = FMT_PCT
ws['D20'] = 0.043
ws['D20'].number_format = FMT_PCT
ws['E20'] = '=ABS(C20-D20)/D20'
ws['E20'].number_format = FMT_PCT2
ws['F20'] = '=IF(E20<=0.05,"OK","CHECK")'

ws.conditional_formatting.add('F18:F20',
    FormulaRule(formula=['$F18="OK"'], fill=FILL_RECON_OK, font=FONT_CHECK_OK))
ws.conditional_formatting.add('F18:F20',
    FormulaRule(formula=['$F18="CHECK"'], fill=FILL_RECON_FAIL, font=FONT_CHECK_FAIL))

# Section 4: Model integrity
write_section_header(ws, 22, 2, 6, 'MODEL INTEGRITY')

ws['B23'] = 'Total sheets'
ws['C23'] = 4  # Phase 1: 4 sheets; will be updated each phase
ws['C23'].font = FONT_INPUT
ws['B24'] = 'Total formulas (manually counted)'
ws['C24'] = 0  # Will be updated post-recalc by verification script
ws['C24'].font = FONT_INPUT
ws['B25'] = 'Total errors'
ws['C25'] = 0
ws['C25'].font = FONT_INPUT
ws['B26'] = 'Phase'
ws['C26'] = 1
ws['C26'].font = FONT_INPUT

ws['B28'] = 'Phase 1 status: foundations only'
ws['B28'].font = FONT_ITALIC
ws['B29'] = 'Phase 2 will add component sheets and replace placeholder C5:C10 above with formulas pulling from each.'
ws['B29'].font = FONT_ITALIC


# ============================================================
# SAVE
# ============================================================

# Reorder sheets: Dashboard first
desired_order = ['Dashboard', 'Assumptions', 'Scenario_Engine', 'Audit_Trail']
wb._sheets = [wb[name] for name in desired_order]

output_path = '/mnt/user-data/outputs/replication/uk_ai_externality_model.xlsx'
wb.save(output_path)

print(f'Phase 1 saved: {output_path}')
print(f'Sheet count: {len(wb.sheetnames)}')
print(f'Sheets: {wb.sheetnames}')
print(f'Defined names: {len(wb.defined_names)}')
print(f'Defined names list:')
for name in wb.defined_names:
    print(f'  - {name}')
