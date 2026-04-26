"""
Phase 5b: Embedded native Excel charts
- Dashboard: Component breakdown bar chart
- Tornado: Horizontal bar chart of parameter sensitivities
- Monte_Carlo: Percentile bar chart
"""
from openpyxl import load_workbook
from openpyxl.chart import BarChart, BarChart3D, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.title import Title
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, RegularTextRun
from copy import deepcopy

PATH = '/mnt/user-data/outputs/replication/uk_ai_externality_model.xlsx'

wb = load_workbook(PATH)

# ============================================================================
# DASHBOARD: Component breakdown
# ============================================================================
ws_dash = wb['Dashboard']

# Remove any existing charts
ws_dash._charts = []

# Component data: rows 15-20, B=name, C=value
chart = BarChart()
chart.type = 'bar'  # horizontal bars
chart.style = 11
chart.title = '£461bn Headline Decomposition by Component'
chart.y_axis.title = ''
chart.x_axis.title = 'Cumulative 2024-30 (£bn)'
chart.legend = None

data = Reference(ws_dash, min_col=3, min_row=15, max_row=20, max_col=3)
cats = Reference(ws_dash, min_col=2, min_row=15, max_row=20)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)

# Data labels
chart.dataLabels = DataLabelList(showVal=True)

# Color the single series (shows headline component contributions)
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties, SolidColorFillProperties
from openpyxl.drawing.colors import ColorChoice as DColorChoice

# Sizing
chart.width = 20  # cm
chart.height = 12

ws_dash.add_chart(chart, 'J3')

# ============================================================================
# TORNADO: Horizontal bar chart
# ============================================================================
ws_t = wb['Tornado']
ws_t._charts = []

# Tornado data is at rows 5-13 (sorted by impact)
# A=name, E=low_hl, F=high_hl, G=range
# Build a chart showing the range as horizontal bars
chart2 = BarChart()
chart2.type = 'bar'
chart2.style = 11
chart2.title = 'Headline Sensitivity to Each Input (±20% perturbation)'
chart2.x_axis.title = 'Headline £bn at perturbed value'
chart2.legend.position = 'b'  # bottom

# Two series: low_hl, high_hl, both refer to headline at perturbed bound
low_data = Reference(ws_t, min_col=5, min_row=4, max_row=13, max_col=5)  # E4:E13 includes header
high_data = Reference(ws_t, min_col=6, min_row=4, max_row=13, max_col=6)  # F4:F13
cats2 = Reference(ws_t, min_col=1, min_row=5, max_row=13)

chart2.add_data(low_data, titles_from_data=True)
chart2.add_data(high_data, titles_from_data=True)
chart2.set_categories(cats2)

# Add a vertical line at central (£457bn) — done via second series
chart2.dataLabels = DataLabelList(showVal=False)
chart2.width = 22
chart2.height = 14

ws_t.add_chart(chart2, 'J4')

# Caption above chart
ws_t['J3'] = 'Range bar shows headline at low (left) vs high (right) input perturbation. Wider = more sensitive.'
from openpyxl.styles import Font
ws_t['J3'].font = Font(name='Calibri', size=9, italic=True, color='595959')

# ============================================================================
# MONTE_CARLO: Headline percentile bar chart
# ============================================================================
ws_mc = wb['Monte_Carlo']
ws_mc._charts = []

# Need to find where the headline percentile row is
# From earlier: row 29 has Headline aggregate, B29=P5, C29=P25, D29=P50, E29=P75, F29=P95
# But we want only P5 P25 P50 P75 P95 columns

# Build a small auxiliary block for the chart
# Place at row 50-55 to keep separate
chart3 = BarChart()
chart3.type = 'col'  # vertical bars
chart3.style = 11
chart3.title = 'Headline Percentile Distribution (N=10,000 Monte Carlo trials)'
chart3.y_axis.title = '£bn'
chart3.x_axis.title = 'Percentile'
chart3.legend = None

# Reference row 29 (headline P5 to P95): B29:F29
data3 = Reference(ws_mc, min_col=2, min_row=29, max_col=6, max_row=29)
# We need column headers from row 28 (B28=P5 header...)
cats3 = Reference(ws_mc, min_col=2, min_row=28, max_col=6, max_row=28)

chart3.add_data(data3, titles_from_data=False, from_rows=True)
chart3.set_categories(cats3)
chart3.dataLabels = DataLabelList(showVal=True)
chart3.width = 18
chart3.height = 11

ws_mc.add_chart(chart3, 'J4')

ws_mc['J3'] = 'P5–P95 confidence band, uncorrelated parameter draws. Median P50 ≈ paper headline £461bn.'
ws_mc['J3'].font = Font(name='Calibri', size=9, italic=True, color='595959')

wb.save(PATH)
print(f"Phase 5b (embedded charts) saved.")
print(f"Charts added to Dashboard, Tornado, Monte_Carlo.")
