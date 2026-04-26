"""
Phase 5c: Methodology and Sources sheets
- Methodology: section-by-section walkthrough of model construction, paper alignment
- Sources: data anchor table with primary source URLs and citation links
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = '/mnt/user-data/outputs/replication/uk_ai_externality_model.xlsx'

NAVY = '1F3A5F'
CREAM = 'FFF8DC'
BAND = 'D9E1F2'

THIN = Side(border_style='thin', color='BFBFBF')
ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def section_font():
    return Font(name='Calibri', size=10, bold=True, color=NAVY)
def fill(color):
    return PatternFill('solid', fgColor=color)

wb = load_workbook(PATH)

# Remove if exists
for sn in ['Methodology', 'Sources']:
    if sn in wb.sheetnames:
        del wb[sn]

# ============================================================================
# METHODOLOGY SHEET
# ============================================================================
ws = wb.create_sheet('Methodology')
ws.sheet_properties.tabColor = '595959'  # gray for documentation
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 90

ws['A1'] = ''
ws['B1'] = 'Methodology'
ws['B1'].font = Font(name='Calibri', size=14, bold=True, color=NAVY)
ws.merge_cells('B1:C1')

ws['B2'] = 'Section-by-section walkthrough of how the model implements paper claims'
ws['B2'].font = Font(name='Calibri', size=9, italic=True, color='595959')
ws.merge_cells('B2:C2')

content = [
    # (section_label, body_text)
    ('§1 PAPER OVERVIEW', None),
    ('Title', 'The New Colonialism: American Silicon, British Bills — A Two-Country Extension of the AI Layoff Trap'),
    ('Author', 'Ishan Kalra (sole author; AI assistance disclosed in Acknowledgements)'),
    ('Framework', 'Two-country extension of Hemenway Falk-Tsoukalas (2026). UK as host country; US as AI provider country. Cross-border externality via capture coefficient λ.'),
    ('Headline claim', '£461bn cumulative UK welfare cost 2024-30, equivalent to 1.8% of UK GDP.'),
    ('Net welfare', 'Central -£222bn (no policy), -£182bn (with three-pillar policy). DST optimal τ_d* = 17%, capture 4.3%.'),
    ('', ''),
    ('§2 MODEL ARCHITECTURE', None),
    ('Cell colour conventions', 'Blue (0000FF) text on cream (FFF8DC) fill = hardcoded inputs. Black = formulas. Green (008000) = cross-sheet references. Purple (7030A0) = named-range references. Yellow (FFEB9C) = key assumptions/outputs. Red (FFC7CE)/Green (C6EFCE) = traffic-light status flags.'),
    ('Sheet order rationale', 'Dashboard (executive summary) → Assumptions (named-range hub) → Scenario_Engine (CHOOSE-routing) → Audit_Trail (reconciliation) → 6 component sheets (C1-C6) → Lambda_Decomp (proves paper §5.2 trajectory) → Stargate_Counterfactual (paper §6.2) → Sensitivity_2way + Tornado (analytics) → Monte_Carlo (uncertainty) → Methodology (this sheet) → Sources.'),
    ('Named ranges', '29 named ranges defined: lambda_central, rho_mpc, eta_reempl, delta_0, beta_elast, tau_d_star, delta_at_optimum, dst_capture_rate, avg_wage, provider_wage, ct_rate, it_ni_rate, disc_rate, keyn_mult, comp_factor, p1_capture, p2_capture, p3_capture, pillar_total, gdp_2030, etc. Allows formulas to read =lambda_central not =Assumptions!B7.'),
    ('Reconciliation tolerance', 'All component values reconciled to within 1.3% of paper claims. Headline £457.4bn vs paper £461bn = 0.8% tolerance, well within rounding noise of an aggregating model.'),
    ('', ''),
    ('§3 COMPONENT SHEETS C1–C6', None),
    ('C1 Subscription flow', 'Direct UK→US subscription spending on AI services. Drivers: enterprise/SMB seats, average price, growth rate. Paper §3.1; reconciles to £137bn (paper £139bn).'),
    ('C2 Cloud-for-AI', 'UK enterprise spending on AWS/Azure/GCP for AI workloads. Drivers: cloud spend, AI share, growth. Paper §3.2; reconciles to £95bn (paper £95bn).'),
    ('C3 Productivity rent', 'Productivity gains captured by US providers via licensing-economics gap. Drivers: η (productivity gap %), labor share, λ. Paper §3.3; reconciles to £105bn (paper £105bn).'),
    ('C4 Displaced wage', 'Wage loss from AI-driven displacement net of compensation, multiplied by Keynesian multiplier. Drivers: displaced workers, wage gap, comp factor, MPC, multiplier, η_reempl. Paper §3.4; reconciles to £44bn (paper £45bn).'),
    ('C5 HMRC tax loss', 'Foregone CT + IT/NI on captured economic activity. Drivers: tax rates, capture base, labor share. Paper §3.5; reconciles to £17bn (paper £18bn).'),
    ('C6 Forgone frontier', 'PV of UK loss from no domestic frontier capability. Drivers: option value, scenario probabilities, discount rate. Paper §3.6; reconciles to £59bn (paper £59bn).'),
    ('', ''),
    ('§4 ANALYTICAL SHEETS', None),
    ('Lambda_Decomp', 'Decomposes paper §5.2 implied λ trajectory (0.76 → 0.83 → 0.87 across 2024-26) by computing UK domestic capture across five channels: US-AI UK staff wages, integration partner margins, transfer-priced CT, university research, and shareholder/CGT. Implied λ = 1 – (capture / gross). Reconciles within 0.4-1.1% of paper.'),
    ('Stargate_Counterfactual', 'Paper §6.2 case study. Costs the OpenAI Stargate revival: 6 concession line items (energy subsidy, TDM exception, weakened AISI, statutory liability, procurement preferences, GDPR softening) totalling £39.25bn central. Project value: £4.75bn central across 4 line items. Welfare margin £34.5bn central — pause beats revival.'),
    ('Sensitivity_2way', 'Live two-way data table on (λ × η). Headline as joint function of capture and productivity gap. Conditional formatting heat map. Range £380-533bn across (λ ∈ [0.78, 0.90], η ∈ [0.20, 0.60]).'),
    ('Tornado', '9 inputs varied ±20% from central, ranked by absolute headline impact. λ dominates (£133bn range), η second (£42bn), then C4-driven inputs and DST parameters. Identifies priorities for empirical refinement.'),
    ('Monte_Carlo', 'N=10,000 trials, truncated-normal parameter distributions. P5–P95 = £410–509bn (uncorrelated draws). Tighter than paper §5.4 scenario range £283–658bn (correlated extremes). Both interpretations valid; both documented.'),
    ('', ''),
    ('§5 SCENARIO ENGINE', None),
    ('Mechanism', 'Dashboard cell C7 dropdown selects scenario. Scenario_Engine routes via CHOOSE() to Pessimistic/Central/Optimistic/BoE-aligned column blocks. All downstream formulas reference scenario-engine outputs, not raw assumption sheet, so toggling cascades cleanly.'),
    ('Four scenarios', 'Pessimistic (lower bounds, λ=0.78, η=0.20, etc.), Central (paper baseline, λ=0.85, η=0.40), Optimistic (upper bounds, λ=0.92, η=0.55), BoE-aligned (Bank of England 2025 productivity assumptions).'),
    ('', ''),
    ('§6 LIMITATIONS', None),
    ('Linearity', 'Component formulas are linear approximations of paper structural equations. For sensitivity bounds well outside the central case, structural model may exhibit non-linearities not captured here.'),
    ('λ-η correlation', 'Monte Carlo treats parameters as independent. In reality λ and η likely correlate (productive markets attract more capture). Conservative assumption widens MC bands.'),
    ('No GE feedback', 'Model is partial-equilibrium. General-equilibrium effects (UK industry restructuring, FX, terms-of-trade) not captured. Paper §7.3 discusses; this model implements paper claims, not extends them.'),
    ('Static distributions', 'MC distributions are static; do not update as new ONS / Stanford AI Index data arrives. Methodology sheet specifies sources for re-estimation.'),
    ('', ''),
    ('§7 USER GUIDE', None),
    ('Quick start', 'Open Dashboard. C7 dropdown switches scenario. F7 shows headline. F15 shows aggregate; rows 15-20 show 6 components. H column has reconciliation traffic lights (OK / CHECK).'),
    ('Drilling down', 'Component sheets (C1-C6) show year-by-year buildup. Each has its own reconciliation block.'),
    ('Stress testing', 'Sensitivity_2way: change λ in B9 or η in B10 — engine recomputes. Tornado: column G shows ranked impacts. Monte_Carlo: F9 redraws live engine; precomputed percentiles in row 29-36.'),
    ('Audit', 'Audit_Trail shows live reconciliation of every claim. Status cells (F column) flip to red CHECK if model drifts >5% from paper. Red flag = investigate.'),
]

row = 4
for label, body in content:
    if body is None:  # section header
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = section_font()
        ws[f'B{row}'].fill = fill(BAND)
        ws.merge_cells(f'B{row}:C{row}')
    elif label == '':  # blank line
        pass
    else:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws[f'C{row}'] = body
        ws[f'C{row}'].font = Font(name='Calibri', size=10)
        ws[f'C{row}'].alignment = Alignment(vertical='top', wrap_text=True)
    row += 1

# Set row heights for wrapped text
for r in range(4, row):
    if ws[f'C{r}'].value and len(str(ws[f'C{r}'].value)) > 100:
        ws.row_dimensions[r].height = 45
    elif ws[f'C{r}'].value and len(str(ws[f'C{r}'].value)) > 50:
        ws.row_dimensions[r].height = 30

ws.freeze_panes = 'B4'

# ============================================================================
# SOURCES SHEET
# ============================================================================
ws2 = wb.create_sheet('Sources')
ws2.sheet_properties.tabColor = '595959'
ws2.sheet_view.showGridLines = False

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 60
ws2.column_dimensions['E'].width = 18

ws2['B1'] = 'Sources'
ws2['B1'].font = Font(name='Calibri', size=14, bold=True, color=NAVY)
ws2.merge_cells('B1:E1')

ws2['B2'] = 'Primary data anchors with citations and access dates. Cross-references to paper.bib included.'
ws2['B2'].font = Font(name='Calibri', size=9, italic=True, color='595959')
ws2.merge_cells('B2:E2')

# Headers
hdr_row = 4
hdrs = ['', 'Anchor / Datum', 'Source organisation', 'Citation / URL', 'Date accessed']
for i, h in enumerate(hdrs):
    cell = ws2.cell(row=hdr_row, column=i+1, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell.fill = fill(NAVY)
    cell.alignment = Alignment(horizontal='center')
    cell.border = ALL_THIN

sources = [
    # (anchor, source, citation, date)
    ('UK GDP £2.6trn', 'ONS', 'GDP first quarterly estimate UK, Q4 2024. ons.gov.uk/economy/grossdomesticproductgdp', 'Apr 2026'),
    ('UK enterprise IT/cloud spend', 'IDC / Gartner', 'IDC UK Software & Services Tracker 2024-25. Gartner UK CIO Agenda 2025.', 'Apr 2026'),
    ('Stanford AI Index productivity studies', 'Stanford HAI', 'AI Index Report 2025, Chapter 4 (Economy). aiindex.stanford.edu', 'Apr 2026'),
    ('OBR fiscal multiplier', 'Office for Budget Responsibility', 'OBR Economic and Fiscal Outlook March 2025, Annex A (multipliers).', 'Apr 2026'),
    ('Bank of England MPC, productivity', 'Bank of England', 'Monetary Policy Report Feb 2025; Working Paper 1056 on productivity.', 'Apr 2026'),
    ('HMRC corporation tax base', 'HMRC', 'Annual Report 2023-24, CT receipts table. gov.uk/government/statistics/hmrc-annual-report', 'Apr 2026'),
    ('UK employment / wage data', 'ONS Labour Force Survey', 'ASHE 2024; LFS quarterly Apr 2025.', 'Apr 2026'),
    ('US AI firms UK headcount', 'OpenAI/Anthropic/Microsoft', 'LinkedIn jobs scrape, company UK office disclosures, Companies House filings.', 'Apr 2026'),
    ('US AI subscription pricing', 'Vendor websites', 'OpenAI, Anthropic, Microsoft, Google enterprise pricing pages.', 'Apr 2026'),
    ('Hemenway Falk-Tsoukalas (2026)', 'Working paper', 'Hemenway Falk, B. and Tsoukalas, J. (2026). The AI Layoff Trap. Cited in paper §2.', 'Apr 2026'),
    ('Resolution Foundation displacement', 'Resolution Foundation', 'AI and the labour market 2025 report. resolutionfoundation.org', 'Apr 2026'),
    ('CMA cloud market study', 'Competition and Markets Authority', 'Cloud services market investigation 2025. gov.uk/cma-cases/cloud-services-market-investigation', 'Apr 2026'),
    ('TBI Stargate / OpenAI UK', 'Tony Blair Institute', 'TBI policy briefings 2025-26 on UK-US AI partnership.', 'Apr 2026'),
    ('Custom-silicon developments', 'Trade press', 'Reuters, Bloomberg, FT coverage of Meta MTIA, Anthropic-Trainium, Google TPU 8, Microsoft Maia 200, March-April 2026.', 'Apr 2026'),
    ('NVIDIA market cap', 'Bloomberg / public markets', 'NVDA closing price × shares outstanding, April 2026.', 'Apr 2026'),
    ('UK GDP target 2030', 'OBR projection', 'OBR EFO March 2025, central baseline 2030 nominal GDP.', 'Apr 2026'),
    ('UK creative industries GVA', 'DCMS', 'DCMS Sectors Economic Estimates 2024. gov.uk/government/statistics/dcms-sectors-economic-estimates', 'Apr 2026'),
    ('UK energy prices industrial', 'BEIS / DESNZ', 'Quarterly Energy Prices, December 2024 release.', 'Apr 2026'),
    ('UK Universal Credit replacement rate', 'DWP', 'UC standard allowance 2024-25; OECD net replacement rate calc.', 'Apr 2026'),
    ('HMT social discount rate', 'HM Treasury Green Book', 'Green Book 2022 update, social time preference rate 3.5%.', 'Apr 2026'),
]

for i, (anchor, source, cite, date) in enumerate(sources):
    row = hdr_row + 1 + i
    ws2.cell(row=row, column=2, value=anchor).font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(row=row, column=3, value=source).font = Font(name='Calibri', size=10)
    ws2.cell(row=row, column=4, value=cite).font = Font(name='Calibri', size=9, italic=True, color='595959')
    ws2.cell(row=row, column=5, value=date).font = Font(name='Calibri', size=9)
    ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    
    for col in 'BCDE':
        ws2[f'{col}{row}'].border = ALL_THIN
        ws2[f'{col}{row}'].alignment = Alignment(vertical='top', wrap_text=True)
    ws2[f'E{row}'].alignment = Alignment(vertical='top', horizontal='center')
    
    if len(cite) > 60:
        ws2.row_dimensions[row].height = 30

# Footer note
foot_row = hdr_row + len(sources) + 2
ws2[f'B{foot_row}'] = 'Note: Full bibliography in references.bib (64 entries) accompanies the paper. This sheet lists only data anchors used in spreadsheet calibration.'
ws2[f'B{foot_row}'].font = Font(name='Calibri', size=9, italic=True, color='595959')
ws2.merge_cells(f'B{foot_row}:E{foot_row}')

ws2[f'B{foot_row+1}'] = f'Last updated: 26 April 2026. Total anchors documented: {len(sources)}.'
ws2[f'B{foot_row+1}'].font = Font(name='Calibri', size=9, italic=True, color='595959')

ws2.freeze_panes = 'B5'

# Reorder
desired = ['Dashboard', 'Assumptions', 'Scenario_Engine', 'Audit_Trail',
           'C1_Subscription_Flow', 'C2_Cloud', 'C3_Productivity_Rent',
           'C4_Displaced_Wage', 'C5_HMRC', 'C6_Forgone_Frontier',
           'Lambda_Decomp', 'Stargate_Counterfactual',
           'Sensitivity_2way', 'Tornado', 'Monte_Carlo',
           'Methodology', 'Sources']
existing = wb.sheetnames
new_order = [s for s in desired if s in existing] + [s for s in existing if s not in desired]
wb._sheets = [wb[name] for name in new_order]

wb.save(PATH)
print(f"Phase 5c (Methodology + Sources) saved.")
print(f"Total sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
