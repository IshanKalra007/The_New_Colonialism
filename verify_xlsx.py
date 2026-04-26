"""
Verify the spreadsheet's computed values match the paper's claims exactly.
Loads with data_only=True to read calculated values from LibreOffice recalc.
"""
from openpyxl import load_workbook

path = '/mnt/user-data/outputs/replication/uk_ai_externality_model.xlsx'
wb = load_workbook(path, data_only=True)

print('=' * 70)
print('VERIFICATION: spreadsheet values vs paper claims')
print('=' * 70)

# ----- Component 1: Direct subscription flow -----
ws = wb['UK_US_Flows']
total_2030_cumulative = ws['B18'].value
component_1_paper = ws['B19'].value
diff_c1 = ws['B20'].value
print(f'\nComponent 1: Direct subscription flow')
print(f'  Spreadsheet cumulative 2023-2030: £{total_2030_cumulative:.1f}bn')
print(f'  Paper claim: £{component_1_paper}bn')
print(f'  Difference: £{diff_c1:.1f}bn')

# Per-year totals for sanity
print(f'\n  Year-by-year totals:')
for i, yr in enumerate(range(2023, 2031)):
    col = chr(ord('B') + i)
    val = ws[f'{col}11'].value
    print(f'    {yr}: £{val:.2f}bn')

# ----- Component 2: Cloud infrastructure -----
ws = wb['Cloud_Infrastructure']
c2_total = ws['I7'].value
c2_paper = ws['B9'].value
print(f'\nComponent 2: Cloud-for-AI flow')
print(f'  Spreadsheet cumulative: £{c2_total:.1f}bn')
print(f'  Paper claim: £{c2_paper}bn')

# ----- Component 3: Productivity rent -----
ws = wb['Productivity_Rent']
c3_total = ws['I7'].value
c3_paper = ws['B9'].value
print(f'\nComponent 3: Productivity rent transfer')
print(f'  Spreadsheet cumulative: £{c3_total:.1f}bn')
print(f'  Paper claim: £{c3_paper}bn')

# ----- Component 4: Displaced wage + multiplier -----
ws = wb['Sectoral_Panel']
c4_total = ws['B18'].value
combined_mult = ws['B16'].value
cumulative_wage = ws['B17'].value
print(f'\nComponent 4: Displaced wage + multiplier')
print(f'  Cumulative wage loss 2023-2030: £{cumulative_wage:.1f}bn')
print(f'  Combined multiplier: {combined_mult}x')
print(f'  Component 4 total: £{c4_total:.1f}bn')
print(f'  Paper claim: £45bn')

# Per-year stock check
print(f'\n  Currently-displaced stock:')
for i, yr in enumerate(range(2023, 2031)):
    col = chr(ord('B') + i)
    val = ws[f'{col}7'].value
    print(f'    {yr}: {val}k')

# Per-year wage loss
print(f'\n  Annual wage loss:')
for i, yr in enumerate(range(2023, 2031)):
    col = chr(ord('B') + i)
    val = ws[f'{col}10'].value
    print(f'    {yr}: £{val:.2f}bn')

# ----- Component 5: HMRC loss -----
ws = wb['HMRC_Loss']
c5_total = ws['I13'].value
c5_paper = ws['B15'].value
c5_2030 = ws['I12'].value
print(f'\nComponent 5: HMRC tax loss')
print(f'  Spreadsheet cumulative: £{c5_total:.1f}bn')
print(f'  Paper claim: £{c5_paper}bn')
print(f'  2030 annual run rate: £{c5_2030:.2f}bn (paper claim £4.7bn)')

# ----- Headline aggregate -----
print(f'\n' + '=' * 70)
print('HEADLINE AGGREGATE')
print('=' * 70)
ws = wb['Headline_Aggregate']
for r in range(4, 10):
    name = ws.cell(row=r, column=1).value
    val = ws.cell(row=r, column=2).value
    print(f'  {name}: £{val:.1f}bn')

total = ws['B11'].value
paper_claim = ws['B12'].value
diff = ws['B13'].value
gdp_pct = ws['B15'].value
print(f'\n  TOTAL: £{total:.1f}bn')
print(f'  Paper claim: £{paper_claim}bn')
print(f'  Reconciliation diff: £{diff:.2f}bn')
print(f'  As % UK GDP: {gdp_pct*100:.2f}%')

# ----- Lambda decomposition -----
print(f'\n' + '=' * 70)
print('LAMBDA DECOMPOSITION CHECK')
print('=' * 70)
ws = wb['Lambda_Decomposition']
for i, yr in enumerate(['2024', '2025', '2026']):
    col = chr(ord('B') + i)
    gross = ws[f'{col}4'].value
    capture = ws[f'{col}13'].value
    lam = ws[f'{col}15'].value
    print(f'  {yr}: gross £{gross:.1f}bn, capture £{capture:.2f}bn, λ = {lam:.3f}')

# ----- DST optimal rate (Proposition 4) -----
print(f'\n' + '=' * 70)
print('PROPOSITION 4 / 5 DST OPTIMUM')
print('=' * 70)
ws = wb['Parameters']
# Find rows for tau_d* and capture rate
# Based on layout: dst_start_row = 17 (after 9 model params + 1 blank + 1 section = row 17)
for r in range(17, 22):
    name = ws.cell(row=r, column=1).value
    val = ws.cell(row=r, column=2).value
    if name and val is not None:
        if isinstance(val, float):
            if 'optimal' in str(name).lower() or 'realised' in str(name).lower() or 'avoidance' in str(name).lower() or 'τ_d' in str(name) or 'δ' in str(name):
                print(f'  {name}: {val:.4f} ({val*100:.2f}%)')

# ----- Counterfactual -----
print(f'\n' + '=' * 70)
print('STARGATE UK COUNTERFACTUAL')
print('=' * 70)
ws = wb['Counterfactual']
print(f'  Total concession bundle (low / high / central):')
print(f'    Low: £{ws["B12"].value:.1f}bn')
print(f'    High: £{ws["C12"].value:.1f}bn')
print(f'    Central: £{ws["D12"].value:.1f}bn')
print(f'  Total project value (low / high / central):')
print(f'    Low: £{ws["B19"].value:.1f}bn')
print(f'    High: £{ws["C19"].value:.1f}bn')
print(f'    Central: £{ws["D19"].value:.1f}bn')
print(f'  Welfare margin (pause beats revival by):')
print(f'    Low: £{ws["B21"].value:.1f}bn')
print(f'    High: £{ws["C21"].value:.1f}bn')
print(f'    Central: £{ws["D21"].value:.1f}bn')

# ----- Scenarios -----
print(f'\n' + '=' * 70)
print('NET WELFARE SCENARIOS')
print('=' * 70)
ws = wb['Scenarios']
labels = ['Pessimistic', 'Central', 'Optimistic', 'BoE']
for i, lbl in enumerate(labels):
    col = chr(ord('B') + i)
    gain = ws[f'{col}4'].value
    rent = ws[f'{col}5'].value
    net_no = ws[f'{col}7'].value
    net_yes = ws[f'{col}10'].value
    print(f'  {lbl}: gain £{gain}, rent £{rent:.0f}, net_no £{net_no:.0f}, net_yes £{net_yes:.0f}')

print(f'\n' + '=' * 70)
print('VERIFICATION COMPLETE')
print('=' * 70)
